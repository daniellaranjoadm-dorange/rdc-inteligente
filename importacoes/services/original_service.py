from __future__ import annotations

import csv
import re
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from django.utils import timezone
from pathlib import Path
from typing import Any

from django.db import transaction
from openpyxl import load_workbook

from acesso.models import RegistroCatraca
from alocacao.models import FuncionarioProjeto
from cadastros.models import AreaLocal, Disciplina, Empresa, Equipe, Funcao, Funcionario, Projeto
from core.choices import StatusImportacaoChoices
from importacoes.models import ImportacaoArquivo, ImportacaoErro
from planejamento.models import AtividadeCronograma


def normalizar_nome_coluna(valor: str) -> str:
    if not valor:
        return ""
    texto = str(valor).strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = texto.encode("ascii", "ignore").decode("ascii")
    texto = re.sub(r"[^a-z0-9]+", "_", texto)
    texto = re.sub(r"_+", "_", texto)
    return texto.strip("_")


def normalizar_texto_busca(valor: str) -> str:
    if not valor:
        return ""
    texto = str(valor).strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = texto.encode("ascii", "ignore").decode("ascii")
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip()


def normalizar_nome_comparacao(valor: str) -> str:
    if not valor:
        return ""
    texto = normalizar_texto_busca(valor)
    for alvo in (" ltda", " ltda.", " sa", ".", "-", "_"):
        texto = texto.replace(alvo, " ")
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip()


def normalizar_cnpj(valor: str) -> str:
    return re.sub(r"\D", "", str(valor or ""))


@dataclass
class ResultadoLinhaImportacao:
    sucesso: bool
    mensagem: str = ""


class LeitorPlanilhaBase:
    COLUNAS_OBRIGATORIAS: tuple[str, ...] = ()
    ALIASES_COLUNAS: dict[str, list[str]] = {}

    def __init__(self, importacao: ImportacaoArquivo):
        self.importacao = importacao

    def processar(self) -> None:
        self.importacao.erros.all().delete()
        self.importacao.status = StatusImportacaoChoices.PROCESSANDO
        self.importacao.iniciado_em = self.importacao.iniciado_em or timezone.now()
        self.importacao.observacoes = ""
        self.importacao.save(update_fields=["status", "iniciado_em", "observacoes"])

        total_erros = 0
        total_sucessos = 0

        try:
            linhas = self.ler_arquivo()
            if not linhas:
                raise ValueError("Arquivo sem registros para processar.")

            with transaction.atomic():
                for indice, linha in enumerate(linhas, start=2):
                    try:
                        self.validar_colunas(linha)
                        resultado = self.processar_linha(linha)
                        if resultado.sucesso:
                            total_sucessos += 1
                        else:
                            total_erros += 1
                            self.registrar_erro(indice, "", resultado.mensagem)
                    except Exception as exc:
                        total_erros += 1
                        self.registrar_erro(indice, "", str(exc))

            self.importacao.status = (
                StatusImportacaoChoices.CONCLUIDO_COM_ERROS
                if total_erros > 0
                else StatusImportacaoChoices.CONCLUIDO
            )
            self.importacao.observacoes = (
                f"Processamento finalizado. Sucessos: {total_sucessos}. "
                f"Erros: {total_erros}."
            )
        except Exception as exc:
            self.importacao.status = StatusImportacaoChoices.ERRO
            self.importacao.observacoes = f"Erro no processamento: {exc}"
        finally:
            self.importacao.finalizado_em = timezone.now()
            self.importacao.save(update_fields=["status", "observacoes", "finalizado_em"])

    def ler_arquivo(self) -> list[dict[str, Any]]:
        caminho = self.importacao.arquivo.path
        sufixo = Path(caminho).suffix.lower()

        if sufixo == ".csv":
            return self._ler_csv(caminho)
        if sufixo in {".xlsx", ".xlsm"}:
            return self._ler_xlsx(caminho)

        raise ValueError("Formato de arquivo não suportado. Use CSV ou XLSX.")

    def _ler_csv(self, caminho: str) -> list[dict[str, Any]]:
        with open(caminho, mode="r", encoding="utf-8-sig", newline="") as arquivo:
            leitor = csv.DictReader(arquivo, delimiter=";")
            linhas = [self.normalizar_chaves(linha) for linha in leitor]

        if not linhas:
            with open(caminho, mode="r", encoding="utf-8-sig", newline="") as arquivo:
                leitor = csv.DictReader(arquivo, delimiter=",")
                linhas = [self.normalizar_chaves(linha) for linha in leitor]

        return linhas

    def _ler_xlsx(self, caminho: str) -> list[dict[str, Any]]:
        wb = load_workbook(caminho, data_only=True)
        ws = wb.active

        cabecalhos = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]

        linhas: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(valor is not None and str(valor).strip() != "" for valor in row):
                continue
            linha = {cabecalhos[idx]: row[idx] if idx < len(row) else None for idx in range(len(cabecalhos))}
            linhas.append(self.normalizar_chaves(linha))
        return linhas

    def normalizar_chaves(self, linha: dict[str, Any]) -> dict[str, Any]:
        nova_linha: dict[str, Any] = {}
        for chave, valor in linha.items():
            chave_norm = normalizar_nome_coluna(chave)
            campo_encontrado = None
            for campo, aliases in self.ALIASES_COLUNAS.items():
                aliases_norm = [normalizar_nome_coluna(alias) for alias in aliases]
                if chave_norm in aliases_norm:
                    campo_encontrado = campo
                    break
            nova_linha[campo_encontrado or chave_norm] = valor
        return nova_linha

    def validar_colunas(self, linha: dict[str, Any]) -> None:
        colunas_ausentes = [col for col in self.COLUNAS_OBRIGATORIAS if col not in linha]
        if colunas_ausentes:
            raise ValueError(f"Colunas obrigatórias ausentes: {', '.join(colunas_ausentes)}")

    def registrar_erro(self, linha: int, campo: str, mensagem: str) -> None:
        ImportacaoErro.objects.create(
            importacao=self.importacao,
            linha=linha,
            campo=campo,
            mensagem=mensagem,
        )

    def processar_linha(self, linha: dict[str, Any]) -> ResultadoLinhaImportacao:
        raise NotImplementedError

    @staticmethod
    def valor_texto(valor: Any) -> str:
        if valor is None:
            return ""
        return str(valor).strip()

    @staticmethod
    def valor_bool(valor: Any, default: bool = True) -> bool:
        if valor is None or str(valor).strip() == "":
            return default
        texto = str(valor).strip().lower()
        return texto in {"1", "true", "sim", "s", "yes", "y"}

    @staticmethod
    def valor_data(valor: Any):
        if valor in (None, ""):
            return None
        if hasattr(valor, "date"):
            return valor.date()
        if isinstance(valor, str):
            for formato in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
                try:
                    return datetime.strptime(valor.strip(), formato).date()
                except ValueError:
                    continue
        raise ValueError(f"Data inválida: {valor}")

    @staticmethod
    def valor_hora(valor: Any):
        if valor in (None, ""):
            return None
        if hasattr(valor, "time"):
            return valor.time()
        if hasattr(valor, "hour") and hasattr(valor, "minute"):
            return valor
        if isinstance(valor, str):
            texto = valor.strip()
            for formato in ("%H:%M", "%H:%M:%S"):
                try:
                    return datetime.strptime(texto, formato).time()
                except ValueError:
                    continue
        raise ValueError(f"Hora inválida: {valor}")

    @staticmethod
    def valor_decimal(valor: Any, default: Decimal = Decimal("0")) -> Decimal:
        if valor in (None, ""):
            return default
        try:
            texto = str(valor).replace(".", "").replace(",", ".")
            return Decimal(texto)
        except Exception as exc:
            raise ValueError(f"Valor numérico inválido: {valor}") from exc


def gerar_codigo_unico_funcao(nome_funcao: str) -> str:
    codigo_base = re.sub(r"[^A-Z0-9]+", "_", normalizar_texto_busca(nome_funcao).upper()).strip("_") or "FUNCAO"
    codigo = codigo_base[:30]
    contador = 1
    while Funcao.objects.filter(codigo=codigo).exists():
        contador += 1
        codigo = f"{codigo_base}_{contador}"[:30]
    return codigo


def buscar_ou_criar_funcao(nome_funcao: str, nome_padrao: str = "SEM FUNÇÃƒO") -> Funcao:
    nome_original = (nome_funcao or "").strip() or nome_padrao
    nome_norm = normalizar_nome_comparacao(nome_original)

    funcoes = list(Funcao.objects.all())
    for funcao in funcoes:
        nome_existente = normalizar_nome_comparacao(funcao.nome)
        if nome_existente == nome_norm or nome_norm in nome_existente or nome_existente in nome_norm:
            return funcao

    return Funcao.objects.create(
        codigo=gerar_codigo_unico_funcao(nome_original),
        nome=nome_original[:120],
        ativa=True,
    )


def buscar_ou_criar_empresa(nome_empresa: str, cnpj: str = "", nome_padrao: str = "EMPRESA NÃƒO INFORMADA") -> Empresa:
    nome_original = (nome_empresa or "").strip() or nome_padrao
    cnpj_normalizado = normalizar_cnpj(cnpj)
    nome_norm = normalizar_nome_comparacao(nome_original)

    if cnpj_normalizado:
        empresa_por_cnpj = Empresa.objects.filter(cnpj=cnpj_normalizado).first()
        if empresa_por_cnpj:
            alteracoes = []
            if nome_original and empresa_por_cnpj.nome != nome_original:
                empresa_por_cnpj.nome = nome_original[:255]
                alteracoes.append("nome")
            if empresa_por_cnpj.cadastro_pendente:
                empresa_por_cnpj.cadastro_pendente = False
                alteracoes.append("cadastro_pendente")
            if empresa_por_cnpj.observacoes:
                empresa_por_cnpj.observacoes = ""
                alteracoes.append("observacoes")
            if alteracoes:
                empresa_por_cnpj.save(update_fields=alteracoes)
            return empresa_por_cnpj

    empresas = list(Empresa.objects.all())
    for empresa in empresas:
        nome_existente = normalizar_nome_comparacao(empresa.nome)
        if nome_existente == nome_norm or nome_norm in nome_existente or nome_existente in nome_norm:
            alteracoes = []
            if cnpj_normalizado and not empresa.cnpj:
                empresa.cnpj = cnpj_normalizado
                empresa.cadastro_pendente = False
                empresa.observacoes = ""
                alteracoes.extend(["cnpj", "cadastro_pendente", "observacoes"])
            if alteracoes:
                empresa.save(update_fields=alteracoes)
            return empresa

    observacoes = "Importada automaticamente sem CNPJ. Completar cadastro." if not cnpj_normalizado else ""
    return Empresa.objects.create(
        nome=nome_original[:255],
        cnpj=cnpj_normalizado if cnpj_normalizado else "",
        ativa=True,
        cadastro_pendente=not bool(cnpj_normalizado),
        observacoes=observacoes,
    )


class ImportadorFuncionarios(LeitorPlanilhaBase):
    COLUNAS_OBRIGATORIAS = ("matricula", "nome")
    ALIASES_COLUNAS = {
        "matricula": ["matricula", "matrícula", "chapa", "id funcionario", "id_funcionario"],
        "nome": ["nome", "nome funcionario", "nome do funcionario", "funcionario", "colaborador"],
        "cpf": ["cpf"],
        "rg": ["rg"],
        "funcao": ["funcao", "função", "cargo", "ocupacao", "ocupAção"],
        "empresa": ["empresa", "empreiteira", "contratada", "relAção social", "relacao social"],
        "cnpj": ["cnpj"],
        "ativo": ["ativo", "status", "situacao", "situAção"],
    }

    def processar_linha(self, linha: dict[str, Any]) -> ResultadoLinhaImportacao:
        matricula = self.valor_texto(linha.get("matricula"))
        nome = self.valor_texto(linha.get("nome"))
        cpf = self.valor_texto(linha.get("cpf"))
        rg = self.valor_texto(linha.get("rg"))
        nome_funcao = self.valor_texto(linha.get("funcao"))
        nome_empresa = self.valor_texto(linha.get("empresa"))
        cnpj_empresa = self.valor_texto(linha.get("cnpj"))
        ativo = self.valor_bool(linha.get("ativo"), default=True)

        if not matricula or not nome:
            return ResultadoLinhaImportacao(False, "Matrícula e nome são obrigatórios.")

        funcionario_existente = Funcionario.objects.filter(matricula=matricula).select_related("funcao", "empresa").first()

        if funcionario_existente and not nome_funcao and funcionario_existente.funcao_id:
            funcao = funcionario_existente.funcao
        else:
            funcao = buscar_ou_criar_funcao(nome_funcao)

        if funcionario_existente and not nome_empresa and funcionario_existente.empresa_id:
            empresa = funcionario_existente.empresa
        else:
            empresa = buscar_ou_criar_empresa(nome_empresa, cnpj_empresa)

        Funcionario.objects.update_or_create(
            matricula=matricula,
            defaults={
                "nome": nome,
                "cpf": cpf,
                "rg": rg,
                "funcao": funcao,
                "empresa": empresa,
                "ativo": ativo,
            },
        )
        return ResultadoLinhaImportacao(True)


class ImportadorCatraca(LeitorPlanilhaBase):
    COLUNAS_OBRIGATORIAS = ("data", "matricula")
    ALIASES_COLUNAS = {
        "data": ["data", "data evento", "dt evento", "data acesso"],
        "matricula": ["matricula", "matrícula", "chapa", "id funcionario", "id_funcionario"],
        "projeto": ["projeto", "obra", "codigo projeto", "código projeto"],
        "entrada_1": ["entrada_1", "entrada", "primeira entrada"],
        "saida_1": ["saida_1", "saída_1", "saida", "saída", "primeira saida", "primeira saída"],
        "entrada_2": ["entrada_2", "segunda entrada"],
        "saida_2": ["saida_2", "segunda saida", "segunda saída"],
        "presente": ["presente", "status acesso"],
        "observacao": ["observacao", "observAção", "obs"],
    }

    def processar_linha(self, linha: dict[str, Any]) -> ResultadoLinhaImportacao:
        data = self.valor_data(linha.get("data"))
        matricula = self.valor_texto(linha.get("matricula"))
        projeto_codigo = self.valor_texto(linha.get("projeto"))
        entrada_1 = self.valor_hora(linha.get("entrada_1"))
        saida_1 = self.valor_hora(linha.get("saida_1"))
        entrada_2 = self.valor_hora(linha.get("entrada_2"))
        saida_2 = self.valor_hora(linha.get("saida_2"))
        presente = self.valor_bool(linha.get("presente"), default=False)
        observacao = self.valor_texto(linha.get("observacao"))

        if not data or not matricula:
            return ResultadoLinhaImportacao(False, "Data e matrícula são obrigatórios.")

        funcionario = Funcionario.objects.filter(matricula=matricula).first()
        projeto = buscar_projeto_por_codigo_ou_nome(projeto_codigo) if projeto_codigo else None

        RegistroCatraca.objects.update_or_create(
            data=data,
            matricula=matricula,
            defaults={
                "funcionario": funcionario,
                "projeto": projeto,
                "entrada_1": entrada_1,
                "saida_1": saida_1,
                "entrada_2": entrada_2,
                "saida_2": saida_2,
                "presente": presente,
                "origem_arquivo": self.importacao.arquivo.name,
                "observacao": observacao,
            },
        )
        return ResultadoLinhaImportacao(True)


class ImportadorAlocacao(LeitorPlanilhaBase):
    COLUNAS_OBRIGATORIAS = ("matricula", "projeto", "disciplina", "data_inicio")
    ALIASES_COLUNAS = {
        "matricula": ["matricula", "matrícula", "chapa", "id funcionario", "id_funcionario"],
        "projeto": ["projeto", "obra", "codigo projeto", "código projeto"],
        "disciplina": ["disciplina"],
        "equipe": ["equipe", "time", "turma"],
        "data_inicio": ["data_inicio", "data inicio", "início", "inicio"],
        "data_fim": ["data_fim", "data fim", "fim"],
        "ativo": ["ativo", "status", "situacao", "situAção"],
    }

    def processar_linha(self, linha: dict[str, Any]) -> ResultadoLinhaImportacao:
        matricula = self.valor_texto(linha.get("matricula"))
        codigo_projeto = self.valor_texto(linha.get("projeto"))
        nome_disciplina = self.valor_texto(linha.get("disciplina"))
        nome_equipe = self.valor_texto(linha.get("equipe"))
        data_inicio = self.valor_data(linha.get("data_inicio"))
        data_fim = self.valor_data(linha.get("data_fim"))
        ativo = self.valor_bool(linha.get("ativo"), default=True)

        funcionario = Funcionario.objects.filter(matricula=matricula).first()
        if not funcionario:
            return ResultadoLinhaImportacao(False, f"Funcionário não encontrado para matrícula {matricula}")

        projeto = buscar_projeto_por_codigo_ou_nome(codigo_projeto)
        if not projeto:
            return ResultadoLinhaImportacao(False, f"Projeto não encontrado: {codigo_projeto}")

        disciplina = buscar_disciplina_por_codigo_ou_nome(nome_disciplina)
        if not disciplina:
            return ResultadoLinhaImportacao(False, f"Disciplina não encontrada: {nome_disciplina}")

        equipe = None
        if nome_equipe:
            equipe = buscar_equipe_por_nome_ou_codigo(nome_equipe, disciplina=disciplina, empresa=funcionario.empresa)
            if not equipe:
                return ResultadoLinhaImportacao(False, f"Equipe não encontrada: {nome_equipe}")

        FuncionarioProjeto.objects.update_or_create(
            funcionario=funcionario,
            projeto=projeto,
            disciplina=disciplina,
            equipe=equipe,
            data_inicio=data_inicio,
            defaults={
                "data_fim": data_fim,
                "ativo": ativo,
            },
        )
        return ResultadoLinhaImportacao(True)



class ImportadorCronograma(LeitorPlanilhaBase):
    COLUNAS_OBRIGATORIAS = ("activity_id", "activity_name", "start", "finish")
    ALIASES_COLUNAS = {
        "item_pai": ["Item Pai", "item_pai"],
        "item": ["Item", "item"],
        "activity_id": ["Activity ID", "activity_id", "atividade id", "id atividade"],
        "activity_name": ["Activity Name", "activity_name", "atividade", "nome atividade"],
        "start": ["Start", "start", "data inicio", "inicio"],
        "finish": ["Finish", "finish", "data fim", "fim"],
        "schedule_percent_complete": ["Schedule % Complete", "schedule_percent_complete", "% completo", "progresso"],
        "budgeted_nonlabor_units": ["Budgeted Nonlabor Units", "budgeted_nonlabor_units", "qtd escopo"],
        "at_completion_nonlabor_units": ["At Completion Nonlabor Units", "at_completion_nonlabor_units"],
    }

    def __init__(self, importacao: ImportacaoArquivo):
        super().__init__(importacao)
        self._projeto = None
        self._disciplinas_cache: dict[str, Disciplina] = {}
        self._areas_cache: dict[str, AreaLocal] = {}

    def processar(self) -> None:
        self.importacao.erros.all().delete()
        self.importacao.status = StatusImportacaoChoices.PROCESSANDO
        self.importacao.iniciado_em = self.importacao.iniciado_em or timezone.now()
        self.importacao.observacoes = ""
        self.importacao.save(update_fields=["status", "iniciado_em", "observacoes"])

        total_erros = 0
        total_sucessos = 0

        try:
            linhas = self.ler_arquivo()
            if not linhas:
                raise ValueError("Arquivo sem registros para processar.")

            self._projeto = self._obter_ou_criar_projeto(linhas)

            with transaction.atomic():
                for indice, linha in enumerate(linhas, start=2):
                    try:
                        self.validar_colunas(linha)
                        resultado = self.processar_linha(linha)
                        if resultado.mensagem == "IGNORAR":
                            continue
                        if resultado.sucesso:
                            total_sucessos += 1
                        else:
                            total_erros += 1
                            self.registrar_erro(indice, "", resultado.mensagem)
                    except Exception as exc:
                        total_erros += 1
                        self.registrar_erro(indice, "", str(exc))

            disciplinas = AtividadeCronograma.objects.filter(projeto=self._projeto).values_list("disciplina__nome", flat=True).distinct()
            areas = AtividadeCronograma.objects.filter(projeto=self._projeto).values_list("area_local__descricao", flat=True).distinct()
            self.importacao.status = (
                StatusImportacaoChoices.CONCLUIDO_COM_ERROS
                if total_erros > 0
                else StatusImportacaoChoices.CONCLUIDO
            )
            self.importacao.observacoes = (
                f"Cronograma importado para o projeto {self._projeto.codigo}. "
                f"Atividades válidas: {total_sucessos}. Erros: {total_erros}. "
                f"Disciplinas: {', '.join(sorted(set(filter(None, disciplinas)))) or 'N/D'}. "
                f"Áreas: {len(set(filter(None, areas)))} cadastradas."
            )
        except Exception as exc:
            self.importacao.status = StatusImportacaoChoices.ERRO
            self.importacao.observacoes = f"Erro no processamento do cronograma: {exc}"
        finally:
            self.importacao.finalizado_em = timezone.now()
            self.importacao.save(update_fields=["status", "observacoes", "finalizado_em"])

    def processar_linha(self, linha: dict[str, Any]) -> ResultadoLinhaImportacao:
        codigo_atividade = self.valor_texto(linha.get("activity_id"))
        if not codigo_atividade:
            return ResultadoLinhaImportacao(True, "IGNORAR")

        descricao = self.valor_texto(linha.get("activity_name"))
        if not descricao:
            return ResultadoLinhaImportacao(False, "Linha com activity_id sem descrição de atividade.")

        data_inicio = self._valor_data_cronograma(linha.get("start"))
        data_fim = self._valor_data_cronograma(linha.get("finish"))
        if not data_inicio or not data_fim:
            return ResultadoLinhaImportacao(False, f"Atividade {codigo_atividade} sem data inicial/final válida.")

        disciplina = self._obter_ou_criar_disciplina(descricao)
        area_local = self._obter_ou_criar_area_local(descricao, disciplina)
        progresso = self._valor_percentual(linha.get("schedule_percent_complete"))
        qtd_escopo = self._obter_qtd_escopo(linha)
        codigo_subatividade = self.valor_texto(linha.get("item"))
        status_planejado = self._status_por_percentual(progresso)

        AtividadeCronograma.objects.update_or_create(
            projeto=self._projeto,
            area_local=area_local,
            disciplina=disciplina,
            codigo_atividade=codigo_atividade,
            defaults={
                "descr_atividade": descricao[:255],
                "codigo_subatividade": codigo_subatividade[:50],
                "descr_subatividade": "",
                "qtd_escopo": qtd_escopo,
                "unidade": "UN",
                "data_inicio": data_inicio,
                "data_fim": data_fim,
                "turno": self._inferir_turno(descricao),
                "status_planejado": status_planejado,
            },
        )
        return ResultadoLinhaImportacao(True)

    def _obter_ou_criar_projeto(self, linhas: list[dict[str, Any]]) -> Projeto:
        codigo = self._extrair_codigo_projeto(linhas)
        projeto = Projeto.objects.filter(codigo__iexact=codigo).first()
        if projeto:
            return projeto
        projeto = Projeto.objects.filter(nome__iexact=codigo).first()
        if projeto:
            return projeto
        return Projeto.objects.create(codigo=codigo[:30], nome=codigo[:255], cliente="N/D", ativo=True)

    def _extrair_codigo_projeto(self, linhas: list[dict[str, Any]]) -> str:
        nome_arquivo = Path(self.importacao.arquivo.name).stem.upper()
        for linha in linhas:
            descricao = self.valor_texto(linha.get("activity_name"))
            if not descricao:
                continue
            match = re.search(r"([A-Z]{2,}\d+[A-Z0-9]*)", descricao.upper())
            if match:
                return match.group(1)[:30]
        fallback = re.sub(r"[^A-Z0-9]+", "", nome_arquivo)[:30]
        return fallback or "PROJETO_RDC"

    def _obter_ou_criar_disciplina(self, descricao: str) -> Disciplina:
        chave = self._classificar_disciplina(descricao)
        if chave in self._disciplinas_cache:
            return self._disciplinas_cache[chave]
        mapeamento = {
            "ELETRICA": ("ELE", "Elétrica"),
            "INSTRUMENTACAO": ("INS", "InstrumentAção"),
            "MECANICA": ("MEC", "MecÃ¢nica"),
            "CIVIL": ("CIV", "Civil"),
            "TUBULACAO": ("TUB", "TubulAção"),
            "COMISSIONAMENTO": ("COM", "Comissionamento"),
        }
        codigo, nome = mapeamento[chave]
        disciplina = Disciplina.objects.filter(codigo=codigo).first() or Disciplina.objects.filter(nome__iexact=nome).first()
        if not disciplina:
            disciplina = Disciplina.objects.create(codigo=codigo, nome=nome, ativo=True)
        self._disciplinas_cache[chave] = disciplina
        return disciplina

    def _classificar_disciplina(self, descricao: str) -> str:
        texto = normalizar_texto_busca(descricao)
        if any(token in texto for token in ["instrument", "connector", "fiber optic", "dio", "loop", "transmitter"]):
            return "INSTRUMENTACAO"
        if any(token in texto for token in ["cable", "busway", "generator", "electrical", "socket", "lighting", "panel", "termination", "cabling"]):
            return "ELETRICA"
        if any(token in texto for token in ["commission", "startup", "energization"]):
            return "COMISSIONAMENTO"
        if any(token in texto for token in ["pipe", "piping", "spool", "line check"]):
            return "TUBULACAO"
        if any(token in texto for token in ["concrete", "civil", "grout", "foundation"]):
            return "CIVIL"
        if any(token in texto for token in ["assembly", "mechanical", "torque", "alignment"]):
            return "MECANICA"
        return "ELETRICA"

    def _obter_ou_criar_area_local(self, descricao: str, disciplina: Disciplina) -> AreaLocal:
        nome_area = self._extrair_area_local(descricao)
        chave = f"{self._projeto.pk}:{nome_area.lower()}"
        if chave in self._areas_cache:
            return self._areas_cache[chave]
        codigo = self._gerar_codigo_area(nome_area)
        area = AreaLocal.objects.filter(projeto=self._projeto, codigo=codigo).first()
        if not area:
            area = AreaLocal.objects.create(
                projeto=self._projeto,
                codigo=codigo,
                descricao=nome_area[:255],
                disciplina_padrao=disciplina,
                ativo=True,
            )
        elif area.disciplina_padrao_id is None:
            area.disciplina_padrao = disciplina
            area.save(update_fields=["disciplina_padrao"])
        self._areas_cache[chave] = area
        return area

    def _extrair_area_local(self, descricao: str) -> str:
        partes = [parte.strip(" -") for parte in re.split(r"\s-\s", descricao) if parte and str(parte).strip()]
        if len(partes) >= 2:
            candidato = partes[-1]
            if len(candidato) >= 3:
                return candidato.upper()[:255]
        return "GERAL"

    def _gerar_codigo_area(self, descricao: str) -> str:
        base = re.sub(r"[^A-Z0-9]+", "_", unicodedata.normalize("NFKD", descricao.upper()).encode("ascii", "ignore").decode("ascii")).strip("_")
        base = re.sub(r"_+", "_", base)[:30]
        return base or "GERAL"

    def _valor_data_cronograma(self, valor: Any):
        if valor in (None, ""):
            return None
        if hasattr(valor, "date"):
            return valor.date()
        texto = str(valor).strip().replace("*", "")
        texto = re.sub(r"\s+[A-Z]$", "", texto)
        for formato in ("%d-%b-%y", "%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(texto, formato).date()
            except ValueError:
                continue
        return self.valor_data(texto)

    def _valor_percentual(self, valor: Any) -> Decimal:
        if valor in (None, ""):
            return Decimal("0")
        texto = str(valor).strip().replace("%", "")
        texto = texto.replace(",", ".")
        try:
            numero = Decimal(texto)
        except Exception:
            return Decimal("0")
        if numero > 1:
            numero = numero / Decimal("100")
        return numero

    def _status_por_percentual(self, percentual: Decimal) -> str:
        if percentual >= Decimal("1"):
            return "Concluída"
        if percentual > Decimal("0"):
            return "Em andamento"
        return "Não iniciada"

    def _obter_qtd_escopo(self, linha: dict[str, Any]) -> Decimal:
        valor = linha.get("at_completion_nonlabor_units")
        if valor in (None, "", 0):
            valor = linha.get("budgeted_nonlabor_units")
        return self.valor_decimal(valor, default=Decimal("0"))

    def _inferir_turno(self, descricao: str) -> str:
        texto = normalizar_texto_busca(descricao)
        if " night " in f" {texto} ":
            return "noite"
        if " morning " in f" {texto} ":
            return "manha"
        if " afternoon " in f" {texto} ":
            return "tarde"
        return "integral"


def importar_cronograma(importacao: ImportacaoArquivo):
    importador = ImportadorCronograma(importacao)
    importador.processar()
    return {"mensagem": importacao.observacoes}


def importar_histograma(importacao: ImportacaoArquivo):
    importacao.status = StatusImportacaoChoices.PROCESSANDO
    importacao.save(update_fields=["status"])
    return {"mensagem": "Stub de importAção de histograma pronto para evolução."}


def buscar_projeto_por_codigo_ou_nome(valor: str) -> Projeto | None:
    termo = normalizar_texto_busca(valor)
    if not termo:
        return None
    for projeto in Projeto.objects.all():
        if normalizar_texto_busca(projeto.codigo) == termo or normalizar_texto_busca(projeto.nome) == termo:
            return projeto
    return Projeto.objects.filter(codigo__iexact=valor).first() or Projeto.objects.filter(nome__iexact=valor).first()


def buscar_disciplina_por_codigo_ou_nome(valor: str) -> Disciplina | None:
    termo = normalizar_texto_busca(valor)
    if not termo:
        return None
    for disciplina in Disciplina.objects.all():
        if normalizar_texto_busca(disciplina.codigo) == termo or normalizar_texto_busca(disciplina.nome) == termo:
            return disciplina
    return Disciplina.objects.filter(codigo__iexact=valor).first() or Disciplina.objects.filter(nome__iexact=valor).first()


def buscar_equipe_por_nome_ou_codigo(valor: str, disciplina: Disciplina | None = None, empresa: Empresa | None = None) -> Equipe | None:
    termo = normalizar_texto_busca(valor)
    if not termo:
        return None
    queryset = Equipe.objects.all()
    if disciplina:
        queryset = queryset.filter(disciplina=disciplina)
    if empresa:
        queryset = queryset.filter(empresa=empresa)
    for equipe in queryset.select_related("disciplina", "empresa"):
        if normalizar_texto_busca(equipe.codigo) == termo or normalizar_texto_busca(equipe.nome) == termo:
            return equipe
    return None



