
from datetime import date, timedelta
from decimal import Decimal
import csv
from io import BytesIO

from django.contrib import messages
from django.db import connection
from django.db.models import Q, Count, Sum
from django.http import FileResponse, Http404, JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse
from core.mixins import AuthenticatedTemplateMixin, RoleRequiredMixin
from core.audit import registrar_auditoria
from django.utils import timezone
from django.views import View
from django.views.generic import (
    CreateView,
    DeleteView,
    DetailView,
    FormView,
    ListView,
    TemplateView,
    UpdateView,
)

from core.mixins import AuthenticatedTemplateMixin
from rdc.forms import (
    RDCAtividadeForm,
    RDCApontamentoForm,
    RDCForm,
    RDCFuncionarioForm,
    RDCMontagemForm,
    RDCValidacaoForm,
)
from rdc.models import RDC, RDCAtividade, RDCFuncionario, RDCValidacao, RDCApontamento
from rdc.services.rdc_service import (
    exportar_rdc_para_modelo_excel,
)

from rdc.services.rdc_montagem_service import (
    montar_rdc_pre_preenchido,
)

from rdc.services.rdc_contexto_service import (
    montar_contexto_montagem_rdc,
    resumo_montagem_rdc,
)

from rdc.services.dashboard_service import (
    build_rdc_dashboard_home_context,
    build_rdc_detail_context,
)
from rdc.services.workflow_service import process_rdc_workflow_action

from rdc.view_helpers import (
    _apply_bool_filter,
    _apply_exact_filter,
    _apply_text_filter,
    _date_shortcuts_for_reference,
    _format_export_value,
    _list_kpis_for_queryset,
    _make_csv_response,
    _montagem_health,
    _quick_filters,
    _resumo_montagem_detalhe,
    _resumo_validacoes,
    _safe_decimal,
    _status_summary_for_queryset,
)


def rdc_tem_schema_fechamento():
    """
    Verifica se os campos de fechamento realmente existem no banco.
    Evita quebrar quando o código já foi atualizado, mas a migration ainda não rodou.
    """
    campos_esperados = {"fechado_em", "fechado_por_id"}

    try:
        tabela = RDC._meta.db_table
        tabelas = connection.introspection.table_names()
        if tabela not in tabelas:
            return False

        with connection.cursor() as cursor:
            descricao = connection.introspection.get_table_description(cursor, tabela)

        nomes_colunas = {col.name for col in descricao}
        return campos_esperados.issubset(nomes_colunas)
    except Exception:
        return False


def _atualizar_validacoes_automaticas(rdc):
    tipos_auto = [
        "funcionario_sem_catraca",
        "duplicidade_apontamento",
        "atividade_fora_cronograma",
    ]
    rdc.validacoes.filter(
        tipo__in=tipos_auto,
        referencia__startswith="AUTO:",
    ).delete()

    if not rdc.atividades.exists():
        RDCValidacao.objects.create(
            rdc=rdc,
            tipo="atividade_fora_cronograma",
            status="bloqueio",
            mensagem="RDC sem atividades cadastradas.",
            referencia="AUTO:SEM_ATIVIDADES",
        )

    for func in rdc.funcionarios.all():
        if not getattr(func, "presente_catraca", False):
            RDCValidacao.objects.create(
                rdc=rdc,
                tipo="funcionario_sem_catraca",
                status="alerta",
                mensagem=f"{func.nome} sem catraca no dia.",
                referencia=f"AUTO:SEM_CATRACA:{func.pk}",
            )

        if not getattr(func, "elegivel", True):
            RDCValidacao.objects.create(
                rdc=rdc,
                tipo="funcionario_sem_catraca",
                status="bloqueio",
                mensagem=f"{func.nome} marcado como não elegível.",
                referencia=f"AUTO:NAO_ELEGIVEL:{func.pk}",
            )

    duplicados = (
        rdc.apontamentos.values("rdc_funcionario_id", "rdc_atividade_id")
        .annotate(total=Count("id"))
        .filter(total__gt=1)
    )
    for item in duplicados:
        RDCValidacao.objects.create(
            rdc=rdc,
            tipo="duplicidade_apontamento",
            status="bloqueio",
            mensagem="Existem apontamentos duplicados para o mesmo funcionário e atividade.",
            referencia=(
                f"AUTO:APONT_DUP:{item['rdc_funcionario_id']}:{item['rdc_atividade_id']}"
            ),
        )



def _get_filtered_rdcs_for_rdo(request):
    qs = RDC.objects.select_related(
        "projeto",
        "area_local",
        "disciplina",
        "supervisor",
    ).all()

    q = request.GET.get("q", "").strip()
    projeto = request.GET.get("projeto", "").strip()
    disciplina = request.GET.get("disciplina", "").strip()
    data = request.GET.get("data", "").strip()

    if data:
        qs = qs.filter(data=data)
    else:
        qs = qs.filter(data=timezone.localdate())

    if projeto:
        qs = qs.filter(projeto_id=projeto)

    if disciplina:
        qs = qs.filter(disciplina_id=disciplina)

    if q:
        qs = _apply_text_filter(
            qs,
            q,
            [
                "projeto__codigo",
                "projeto__nome",
                "disciplina__nome",
                "area_local__descricao",
                "observacoes",
            ],
        )

    return qs.order_by("projeto__codigo", "disciplina__nome", "area_local__descricao", "id")


def _resolve_rdo_scope(request):
    rdcs_filtrados = _get_filtered_rdcs_for_rdo(request)
    ids_query = [i for i in request.GET.getlist("ids") if str(i).strip()]

    if ids_query:
        ids_int = []
        for item in ids_query:
            try:
                ids_int.append(int(item))
            except Exception:
                continue
        rdcs_selecionados = rdcs_filtrados.filter(id__in=ids_int)
        ids_selecionados = list(rdcs_selecionados.values_list("id", flat=True))
    else:
        rdcs_selecionados = rdcs_filtrados
        ids_selecionados = list(rdcs_filtrados.values_list("id", flat=True))

    atividades = RDCAtividade.objects.select_related(
        "rdc",
        "rdc__projeto",
        "rdc__disciplina",
        "rdc__area_local",
    ).filter(rdc_id__in=ids_selecionados)

    funcionarios = RDCFuncionario.objects.select_related(
        "rdc",
        "rdc__projeto",
        "rdc__disciplina",
        "equipe",
        "funcao",
    ).filter(rdc_id__in=ids_selecionados)

    apontamentos = RDCApontamento.objects.select_related(
        "rdc",
        "rdc__projeto",
        "rdc_funcionario",
        "rdc_atividade",
    ).filter(rdc_id__in=ids_selecionados)

    validacoes = RDCValidacao.objects.select_related(
        "rdc",
        "rdc__projeto",
    ).filter(rdc_id__in=ids_selecionados)

    observacoes_consolidadas = [
        obs
        for obs in rdcs_selecionados.exclude(observacoes__isnull=True)
        .exclude(observacoes__exact="")
        .values_list("observacoes", flat=True)
    ]

    resumo_disciplina = list(
        atividades.values("rdc__disciplina__nome")
        .annotate(
            atividades=Count("id"),
            escopo=Sum("qtd_escopo"),
            executado=Sum("qtd_executada"),
        )
        .order_by("rdc__disciplina__nome")
    )

    totais = {
        "rdcs": rdcs_selecionados.count(),
        "atividades": atividades.count(),
        "funcionarios": funcionarios.count(),
        "apontamentos": apontamentos.count(),
        "hh_total": funcionarios.aggregate(total=Sum("hh_total"))["total"] or Decimal("0.00"),
        "alertas": validacoes.filter(status="alerta").count(),
        "bloqueios": validacoes.filter(status="bloqueio").count(),
    }

    return {
        "rdcs_filtrados": rdcs_filtrados,
        "rdcs_selecionados": rdcs_selecionados,
        "ids_selecionados": ids_selecionados,
        "atividades": atividades,
        "funcionarios": funcionarios,
        "apontamentos": apontamentos,
        "validacoes": validacoes,
        "observacoes_consolidadas": observacoes_consolidadas,
        "resumo_disciplina": resumo_disciplina,
        "totais": totais,
    }


def _rdo_table_rows(scope):
    rows = []
    rdcs = list(scope["rdcs_selecionados"])
    atividade_map = {
        item["rdc_id"]: item
        for item in scope["atividades"]
        .values("rdc_id")
        .annotate(
            atividades=Count("id"),
            escopo=Sum("qtd_escopo"),
            executado=Sum("qtd_executada"),
        )
    }
    funcionarios_map = {
        item["rdc_id"]: item
        for item in scope["funcionarios"]
        .values("rdc_id")
        .annotate(
            funcionarios=Count("id"),
            hh_total=Sum("hh_total"),
            presentes=Count("id", filter=Q(presente_catraca=True)),
        )
    }

    for r in rdcs:
        a = atividade_map.get(r.id, {})
        f = funcionarios_map.get(r.id, {})
        rows.append(
            [
                r.id,
                r.data.strftime("%d/%m/%Y") if r.data else "",
                getattr(r.projeto, "codigo", ""),
                getattr(r.disciplina, "nome", ""),
                getattr(r.area_local, "descricao", ""),
                r.get_turno_display() if hasattr(r, "get_turno_display") else r.turno,
                getattr(r, "get_status_display", lambda: r.status)(),
                a.get("atividades", 0) or 0,
                a.get("escopo", Decimal("0.00")) or Decimal("0.00"),
                a.get("executado", Decimal("0.00")) or Decimal("0.00"),
                f.get("funcionarios", 0) or 0,
                f.get("presentes", 0) or 0,
                f.get("hh_total", Decimal("0.00")) or Decimal("0.00"),
                (r.observacoes or "").strip(),
            ]
        )
    return rows


def _make_xlsx_response(
    filename,
    headers,
    rows,
    sheet_name="RDO",
    title=None,
    subtitle=None,
    summary_rows=None,
    extra_sheets=None,
):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or "Relatório")[:31]

    title = title or sheet_name or "Relatório"
    total_cols = max(2, len(headers))
    current_row = 1
    thin = Side(style="thin", color="D1D5DB")

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
    c = ws.cell(row=current_row, column=1, value=title)
    c.font = Font(bold=True, size=16, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="163A63")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[current_row].height = 24
    current_row += 1

    if subtitle:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
        c = ws.cell(row=current_row, column=1, value=subtitle)
        c.font = Font(italic=True, size=10, color="4B5563")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[current_row].height = 34
        current_row += 2
    else:
        current_row += 1

    if summary_rows:
        ws.cell(row=current_row, column=1, value="Resumo executivo").font = Font(bold=True, size=12, color="163A63")
        current_row += 1
        for indicador, valor in summary_rows:
            c1 = ws.cell(row=current_row, column=1, value=indicador)
            c2 = ws.cell(row=current_row, column=2, value=_format_export_value(valor))
            c1.font = Font(bold=True)
            for c in (c1, c2):
                c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                c.alignment = Alignment(vertical="center")
            current_row += 1
        current_row += 2

    header_row = current_row
    for col_idx, header in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=header)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4F7D")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, row in enumerate(rows, start=header_row + 1):
        for col_idx, value in enumerate(row, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=_format_export_value(value))
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = f"A{header_row + 1}"
    if rows:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row + len(rows)}"

    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        candidates = [len(str(headers[col_idx - 1]))]
        for row in rows:
            if len(row) >= col_idx:
                candidates.append(len(_format_export_value(row[col_idx - 1])))
        ws.column_dimensions[col_letter].width = min(max(max(candidates) + 3, 14), 42)

    if extra_sheets:
        for sheet in extra_sheets:
            extra_headers = sheet.get("headers") or []
            extra_rows = sheet.get("rows") or []
            extra_ws = wb.create_sheet(title=(sheet.get("name") or "Dados")[:31])

            title_text = sheet.get("title") or sheet.get("name") or "Dados"
            extra_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(2, len(extra_headers)))
            c = extra_ws.cell(row=1, column=1, value=title_text)
            c.font = Font(bold=True, size=14, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="163A63")
            c.alignment = Alignment(horizontal="center", vertical="center")
            row0 = 3

            for col_idx, header in enumerate(extra_headers, start=1):
                c = extra_ws.cell(row=row0, column=col_idx, value=header)
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="1F4F7D")
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for row_idx, row in enumerate(extra_rows, start=row0 + 1):
                for col_idx, value in enumerate(row, start=1):
                    c = extra_ws.cell(row=row_idx, column=col_idx, value=_format_export_value(value))
                    c.alignment = Alignment(vertical="top", wrap_text=True)
                    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            extra_ws.freeze_panes = f"A{row0 + 1}"
            if extra_rows and extra_headers:
                extra_ws.auto_filter.ref = f"A{row0}:{get_column_letter(len(extra_headers))}{row0 + len(extra_rows)}"

            for col_idx in range(1, len(extra_headers) + 1):
                col_letter = get_column_letter(col_idx)
                candidates = [len(str(extra_headers[col_idx - 1]))]
                for row in extra_rows:
                    if len(row) >= col_idx:
                        candidates.append(len(_format_export_value(row[col_idx - 1])))
                extra_ws.column_dimensions[col_letter].width = min(max(max(candidates) + 3, 14), 42)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _make_pdf_response(
    filename,
    title,
    headers,
    rows,
    subtitle=None,
    summary_rows=None,
):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "RptTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        alignment=1,
        textColor=colors.HexColor("#111827"),
        spaceAfter=4,
    )
    subtitle_style = ParagraphStyle(
        "RptSubtitle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        alignment=1,
        textColor=colors.HexColor("#4B5563"),
        spaceAfter=10,
    )
    section_style = ParagraphStyle(
        "RptSection",
        parent=styles["Heading3"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#163A63"),
        spaceAfter=6,
    )

    story = [Paragraph(title, title_style)]
    if subtitle:
        story.append(Paragraph(subtitle, subtitle_style))
    story.append(Spacer(1, 4))

    if summary_rows:
        story.append(Paragraph("Resumo executivo", section_style))
        summary_table_data = [["Indicador", "Valor"]] + [[str(k), _format_export_value(v)] for k, v in summary_rows]
        summary_table = Table(summary_table_data, colWidths=[95 * mm, 40 * mm], hAlign="LEFT")
        summary_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4F7D")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        story.append(summary_table)
        story.append(Spacer(1, 10))

    if rows:
        story.append(Paragraph("Detalhamento gerencial", section_style))
        table_data = [headers] + [[_format_export_value(v) for v in row] for row in rows]
        page_width = landscape(A4)[0] - doc.leftMargin - doc.rightMargin
        col_width = page_width / max(1, len(headers))
        table = Table(table_data, repeatRows=1, colWidths=[col_width] * len(headers))
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4F7D")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 5),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(table)

    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

class RDCListView(AuthenticatedTemplateMixin, ListView):
    model = RDC
    template_name = "rdc/lista_rdc.html"
    context_object_name = "rdcs"
    paginate_by = 20
    ordering = ["-data", "-id"]

    def _base_queryset(self):
        return (
            super()
            .get_queryset()
            .select_related(
                "projeto",
                "area_local",
                "disciplina",
                "supervisor",
            )
        )

    def get_queryset(self):
        qs = self._base_queryset()
        q = self.request.GET.get("q")
        status = self.request.GET.get("status")
        turno = self.request.GET.get("turno")
        data_ini = (self.request.GET.get("data_ini") or "").strip()
        data_fim = (self.request.GET.get("data_fim") or "").strip()
        projeto = (self.request.GET.get("projeto") or "").strip()
        disciplina = (self.request.GET.get("disciplina") or "").strip()
        quick = (self.request.GET.get("quick") or "").strip()

        if quick:
            quick_filters = _quick_filters().get(quick)
            if quick_filters:
                data_ini = quick_filters.get("data_ini", data_ini)
                data_fim = quick_filters.get("data_fim", data_fim)

        qs = _apply_text_filter(
            qs,
            q,
            [
                "projeto__codigo",
                "projeto__nome",
                "area_local__codigo",
                "area_local__descricao",
                "disciplina__nome",
                "supervisor__username",
                "observacoes",
            ],
        )
        qs = _apply_exact_filter(qs, status, "status")
        qs = _apply_exact_filter(qs, turno, "turno")
        qs = _apply_exact_filter(qs, projeto, "projeto_id")
        qs = _apply_exact_filter(qs, disciplina, "disciplina_id")

        if data_ini:
            qs = qs.filter(data__gte=data_ini)
        if data_fim:
            qs = qs.filter(data__lte=data_fim)

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        queryset = self.object_list
        page_qs = list(context.get("rdcs", []))
        status_summary = _status_summary_for_queryset(queryset)

        context["filter_values"] = {
            "q": self.request.GET.get("q", ""),
            "status": self.request.GET.get("status", ""),
            "turno": self.request.GET.get("turno", ""),
            "data_ini": self.request.GET.get("data_ini", ""),
            "data_fim": self.request.GET.get("data_fim", ""),
            "projeto": self.request.GET.get("projeto", ""),
            "disciplina": self.request.GET.get("disciplina", ""),
            "quick": self.request.GET.get("quick", ""),
        }
        context["quick_filters"] = _quick_filters()
        context["kpis"] = _list_kpis_for_queryset(queryset, page_qs=page_qs)
        context["status_summary"] = status_summary
        context["view_mode"] = self.request.GET.get("view", "cards")
        context["has_active_filters"] = any(
            bool(v)
            for v in context["filter_values"].values()
            if isinstance(v, str)
        )
        context["projetos_filtro"] = (
            RDC.objects.select_related("projeto")
            .values("projeto_id", "projeto__codigo", "projeto__nome")
            .distinct()
            .order_by("projeto__codigo")
        )
        context["disciplinas_filtro"] = (
            RDC.objects.select_related("disciplina")
            .values("disciplina_id", "disciplina__nome")
            .distinct()
            .order_by("disciplina__nome")
        )
        return context


class RDCMontagemView(AuthenticatedTemplateMixin, RoleRequiredMixin, FormView):
    allowed_roles = ["admin", "supervisor"]
    template_name = "rdc/novo_rdc.html"
    form_class = RDCMontagemForm

    def _data_referencia(self):
        raw = self.request.POST.get("data") or self.request.GET.get("data")
        if raw:
            try:
                return datetime.strptime(raw, "%Y-%m-%d").date()
            except Exception:
                pass
        return timezone.localdate()

    def _contexto_guiado(self):
        if not hasattr(self, "_contexto_guiado_cache"):
            self._contexto_guiado_cache = montar_contexto_montagem_rdc(
                self.request.user,
                self._data_referencia(),
            )
        return self._contexto_guiado_cache

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["contexto_guiado"] = self._contexto_guiado()
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        contexto = self._contexto_guiado()
        resumo = resumo_montagem_rdc(contexto)
        data_ref = self._data_referencia()

        context["contexto_guiado"] = contexto
        context["resumo_guiado"] = resumo
        context["semana_planejamento"] = contexto.get("semana")
        context["tem_contexto_minimo"] = contexto.get("tem_contexto_minimo", False)
        context["data_referencia"] = data_ref
        context["date_shortcuts"] = _date_shortcuts_for_reference(data_ref)
        context["montagem_health"] = _montagem_health(contexto, resumo)
        context["atalhos_admin"] = [
            {"label": "Perfis operacionais", "url": "/admin/rdc/perfiloperacionalusuario/"},
            {"label": "ProgramAção semanal", "url": "/admin/rdc/programacaosemanal/"},
        ]
        return context

    def form_valid(self, form):
        contexto = self._contexto_guiado()
        resumo = resumo_montagem_rdc(contexto)
        health = _montagem_health(contexto, resumo)

        forcar_montagem = (self.request.POST.get("forcar_montagem") or "").strip() in {"1", "true", "on"}
        justificativa_excecao = (self.request.POST.get("justificativa_excecao") or "").strip()

        if health.get("score", 0) <= 1:
            if not forcar_montagem:
                form.add_error(
                    None,
                    "Montagem bloqueada: contexto insuficiente para criar o RDC com segurança."
                )
                messages.error(
                    self.request,
                    "Montagem bloqueada. Revise equipe, atividades sugeridas, semana e contexto operacional antes de continuar."
                )
                return self.form_invalid(form)

            if not justificativa_excecao:
                form.add_error(
                    None,
                    "Informe a justificativa para realizar uma montagem excepcional."
                )
                messages.error(
                    self.request,
                    "A justificativa é obrigatória para montagem excepcional."
                )
                return self.form_invalid(form)

        rdc = montar_rdc_pre_preenchido(
            projeto_id=form.cleaned_data["projeto"].id,
            area_local_id=form.cleaned_data["area_local"].id,
            disciplina_id=form.cleaned_data["disciplina"].id,
            data=form.cleaned_data["data"],
            turno=form.cleaned_data["turno"],
            user=self.request.user,
        )

        if health.get("score", 0) <= 1 and forcar_montagem and justificativa_excecao:
            prefixo = (
                "[MONTAGEM EXCEPCIONAL] "
                f"Usuário: {self.request.user} | "
                f"Score: {health.get('score', 0)}/4 | "
                f"Justificativa: {justificativa_excecao}"
            )
            rdc.observacoes = f"{prefixo}\n{rdc.observacoes}".strip()
            rdc.save(update_fields=["observacoes", "updated_at", "sync_updated_at"])
            messages.warning(
                self.request,
                "RDC criado por exceção controlada. A justificativa foi registrada."
            )
        else:
            messages.success(self.request, "RDC pré-preenchido montado com sucesso.")

        return redirect(reverse("rdc-detail", kwargs={"pk": rdc.pk}))


class RDCDetailView(AuthenticatedTemplateMixin, DetailView):
    model = RDC
    template_name = "rdc/detalhe_rdc.html"
    context_object_name = "rdc"

    def _filtered_sets(self, rdc):
        atividades = _apply_text_filter(
            rdc.atividades.all(),
            self.request.GET.get("afq"),
            [
                "codigo_atividade",
                "descr_atividade",
                "codigo_subatividade",
                "descr_subatividade",
            ],
        )
        atividades = _apply_bool_filter(
            atividades,
            self.request.GET.get("astatus"),
            "ativa_no_dia",
        )
        atividades = _apply_exact_filter(
            atividades,
            self.request.GET.get("aorigem"),
            "origem",
        )

        funcionarios = rdc.funcionarios.select_related("equipe", "funcao", "funcionario").all()
        funcionarios = _apply_text_filter(
            funcionarios,
            self.request.GET.get("ffq"),
            ["matricula", "nome", "funcao__nome", "equipe__nome", "motivo_bloqueio"],
        )
        funcionarios = _apply_bool_filter(
            funcionarios,
            self.request.GET.get("felegivel"),
            "elegivel",
        )
        funcionarios = _apply_bool_filter(
            funcionarios,
            self.request.GET.get("fcatraca"),
            "presente_catraca",
        )

        apontamentos = rdc.apontamentos.select_related(
            "rdc_funcionario",
            "rdc_atividade",
        ).all()
        apontamentos = _apply_text_filter(
            apontamentos,
            self.request.GET.get("apq"),
            [
                "rdc_funcionario__nome",
                "rdc_funcionario__matricula",
                "rdc_atividade__codigo_atividade",
                "rdc_atividade__descr_atividade",
                "observacao",
            ],
        )

        validacoes = _apply_text_filter(
            rdc.validacoes.all(),
            self.request.GET.get("vq"),
            ["tipo", "status", "mensagem", "referencia"],
        )
        validacoes = _apply_exact_filter(
            validacoes,
            self.request.GET.get("vstatus"),
            "status",
        )

        return atividades, funcionarios, apontamentos, validacoes

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        rdc = context["rdc"]

        atividades, funcionarios_qs, apontamentos_qs, validacoes_qs = self._filtered_sets(rdc)
        context["atividades_filtradas"] = atividades
        context["funcionarios_filtrados"] = funcionarios_qs
        context["apontamentos_filtrados"] = apontamentos_qs
        context["validacoes_filtradas"] = validacoes_qs
        context["filter_values"] = {
            k: self.request.GET.get(k, "")
            for k in [
                "afq",
                "astatus",
                "aorigem",
                "ffq",
                "felegivel",
                "fcatraca",
                "apq",
                "vq",
                "vstatus",
            ]
        }
        context["schema_fechamento_ok"] = rdc_tem_schema_fechamento()
        context.update(build_rdc_detail_context(rdc, user=self.request.user))
        context["resumo_montagem"] = _resumo_montagem_detalhe(rdc)

        context["montagem_excepcional"] = False
        context["justificativa_excepcional"] = ""

        if rdc.observacoes and rdc.observacoes.startswith("[MONTAGEM EXCEPCIONAL]"):
            context["montagem_excepcional"] = True
            context["justificativa_excepcional"] = rdc.observacoes

        return context


class RDCConsolidadoView(AuthenticatedTemplateMixin, TemplateView):
    template_name = "rdc/painel_consolidado.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        rdcs = RDC.objects.select_related(
            "projeto",
            "area_local",
            "disciplina",
            "supervisor",
        ).all()
        atividades = RDCAtividade.objects.select_related(
            "rdc",
            "rdc__projeto",
            "rdc__disciplina",
            "rdc__area_local",
        ).all()
        funcionarios = RDCFuncionario.objects.select_related(
            "rdc",
            "rdc__projeto",
            "rdc__disciplina",
            "equipe",
            "funcao",
        ).all()
        apontamentos = RDCApontamento.objects.select_related(
            "rdc",
            "rdc__projeto",
            "rdc_funcionario",
            "rdc_atividade",
        ).all()
        validacoes = RDCValidacao.objects.select_related("rdc", "rdc__projeto").all()

        q = self.request.GET.get("q", "")
        projeto = self.request.GET.get("projeto", "")
        disciplina = self.request.GET.get("disciplina", "")
        data_ini = self.request.GET.get("data_ini", "")
        data_fim = self.request.GET.get("data_fim", "")
        aba = self.request.GET.get("aba", "resumo")

        if projeto:
            rdcs = rdcs.filter(projeto_id=projeto)
            atividades = atividades.filter(rdc__projeto_id=projeto)
            funcionarios = funcionarios.filter(rdc__projeto_id=projeto)
            apontamentos = apontamentos.filter(rdc__projeto_id=projeto)
            validacoes = validacoes.filter(rdc__projeto_id=projeto)

        if disciplina:
            rdcs = rdcs.filter(disciplina_id=disciplina)
            atividades = atividades.filter(rdc__disciplina_id=disciplina)
            funcionarios = funcionarios.filter(rdc__disciplina_id=disciplina)
            apontamentos = apontamentos.filter(rdc__disciplina_id=disciplina)
            validacoes = validacoes.filter(rdc__disciplina_id=disciplina)

        if data_ini:
            rdcs = rdcs.filter(data__gte=data_ini)
            atividades = atividades.filter(rdc__data__gte=data_ini)
            funcionarios = funcionarios.filter(rdc__data__gte=data_ini)
            apontamentos = apontamentos.filter(rdc__data__gte=data_ini)
            validacoes = validacoes.filter(rdc__data__gte=data_ini)

        if data_fim:
            rdcs = rdcs.filter(data__lte=data_fim)
            atividades = atividades.filter(rdc__data__lte=data_fim)
            funcionarios = funcionarios.filter(rdc__data__lte=data_fim)
            apontamentos = apontamentos.filter(rdc__data__lte=data_fim)
            validacoes = validacoes.filter(rdc__data__lte=data_fim)

        if q:
            rdcs = _apply_text_filter(
                rdcs,
                q,
                ["projeto__codigo", "area_local__descricao", "disciplina__nome", "observacoes"],
            )
            atividades = _apply_text_filter(
                atividades,
                q,
                [
                    "codigo_atividade",
                    "descr_atividade",
                    "codigo_subatividade",
                    "descr_subatividade",
                    "rdc__projeto__codigo",
                ],
            )
            funcionarios = _apply_text_filter(
                funcionarios,
                q,
                [
                    "matricula",
                    "nome",
                    "funcao__nome",
                    "equipe__nome",
                    "rdc__projeto__codigo",
                ],
            )
            apontamentos = _apply_text_filter(
                apontamentos,
                q,
                [
                    "rdc_funcionario__nome",
                    "rdc_funcionario__matricula",
                    "rdc_atividade__codigo_atividade",
                    "rdc_atividade__descr_atividade",
                ],
            )
            validacoes = _apply_text_filter(
                validacoes,
                q,
                ["tipo", "status", "mensagem", "referencia", "rdc__projeto__codigo"],
            )

        from cadastros.models import Projeto, Disciplina

        context["projetos"] = Projeto.objects.filter(ativo=True).order_by("codigo")
        context["disciplinas"] = Disciplina.objects.filter(ativo=True).order_by("nome")
        context["rdcs"] = rdcs[:300]
        context["atividades"] = atividades[:500]
        context["funcionarios"] = funcionarios[:500]
        context["apontamentos"] = apontamentos[:500]
        context["validacoes"] = validacoes[:500]
        context["filter_values"] = {
            "q": q,
            "projeto": projeto,
            "disciplina": disciplina,
            "data_ini": data_ini,
            "data_fim": data_fim,
        }
        context["aba"] = aba
        context["quick_filters"] = _quick_filters()
        context["schema_fechamento_ok"] = rdc_tem_schema_fechamento()
        context["totais"] = {
            "rdcs": rdcs.count(),
            "atividades": atividades.count(),
            "funcionarios": funcionarios.count(),
            "apontamentos": apontamentos.count(),
            "validacoes": validacoes.count(),
            "hh_total": funcionarios.aggregate(total=Sum("hh_total"))["total"] or Decimal("0.00"),
            "bloqueios": validacoes.filter(status="bloqueio").count(),
            "alertas": validacoes.filter(status="alerta").count(),
            "fechados": rdcs.filter(status="fechado").count(),
            "sem_presenca": funcionarios.filter(presente_catraca=False).count(),
        }

        context["dashboard"] = {
            "top_atividades": list(
                atividades.values("codigo_atividade")
                .annotate(total_exec=Sum("qtd_executada"))
                .order_by("-total_exec", "codigo_atividade")[:10]
            ),
            "top_funcionarios": list(
                funcionarios.values("nome")
                .annotate(hh_total=Sum("hh_total"))
                .order_by("-hh_total", "nome")[:10]
            ),
            "hh_por_dia": list(
                funcionarios.values("rdc__data")
                .annotate(total=Sum("hh_total"))
                .order_by("rdc__data")
            ),
            "hh_por_disciplina": list(
                funcionarios.values("rdc__disciplina__nome")
                .annotate(total=Sum("hh_total"))
                .order_by("-total", "rdc__disciplina__nome")
            ),
            "validacoes_por_tipo": list(
                validacoes.values("tipo")
                .annotate(total=Count("id"))
                .order_by("-total", "tipo")
            ),
        }

        atividades_previsto = atividades.aggregate(total=Sum("qtd_escopo"))["total"] or Decimal("0.00")
        atividades_realizado = atividades.aggregate(total=Sum("qtd_executada"))["total"] or Decimal("0.00")
        hh_realizado = funcionarios.aggregate(total=Sum("hh_total"))["total"] or Decimal("0.00")
        hh_previsto = hh_realizado
        equipe_realizada = funcionarios.count()
        presentes = funcionarios.filter(presente_catraca=True).count()

        percentual_escopo = Decimal("0.00")
        if atividades_previsto:
            percentual_escopo = (Decimal(str(atividades_realizado)) / Decimal(str(atividades_previsto))) * Decimal("100")

        percentual_hh = Decimal("100.00") if hh_previsto else Decimal("0.00")
        if hh_previsto:
            percentual_hh = (Decimal(str(hh_realizado)) / Decimal(str(hh_previsto))) * Decimal("100")

        context["previsto_realizado"] = {
            "atividades_previsto": atividades_previsto,
            "atividades_realizado": atividades_realizado,
            "percentual_escopo": percentual_escopo,
            "escopo_nivel": "success" if percentual_escopo >= 90 else "warning" if percentual_escopo >= 70 else "danger",
            "escopo_texto": "Ritmo aderente ao planejado." if percentual_escopo >= 90 else "Abaixo do esperado.",
            "hh_previsto": hh_previsto,
            "hh_realizado": hh_realizado,
            "percentual_hh": percentual_hh,
            "hh_nivel": "success" if percentual_hh >= 90 else "warning" if percentual_hh >= 70 else "danger",
            "hh_texto": "Consumo de HH compatível." if percentual_hh >= 90 else "Monitorar consumo de HH.",
            "equipe_prevista": equipe_realizada,
            "equipe_realizada": equipe_realizada,
            "presentes": presentes,
            "lancados": equipe_realizada,
        }

        conformidade = Decimal("100.00")
        total_validacoes = validacoes.count()
        if total_validacoes:
            conformidade = Decimal("100.00") - (
                (Decimal(str(validacoes.filter(status="bloqueio").count())) / Decimal(str(total_validacoes)))
                * Decimal("100")
            )

        context["executivos"] = {
            "conformidade": {
                "nivel": "success" if conformidade >= 90 else "warning" if conformidade >= 70 else "danger",
                "rotulo": "Conformidade",
                "valor": f"{conformidade.quantize(Decimal('0.1'))}",
                "sufixo": "%",
                "descricao": "Semáforo gerencial das validaçÃµes.",
            },
            "prontos": {
                "nivel": "success",
                "rotulo": "RDCs prontos",
                "valor": rdcs.filter(status__in=["aprovado", "fechado"]).count(),
                "descricao": "Aprovados ou fechados no filtro.",
            },
            "criticos": {
                "nivel": "danger" if context["totais"]["bloqueios"] else "secondary",
                "rotulo": "RDCs críticos",
                "valor": validacoes.filter(status="bloqueio").values("rdc_id").distinct().count(),
                "descricao": "Com bloqueios ativos.",
            },
            "sem_presenca": {
                "nivel": "warning" if context["totais"]["sem_presenca"] else "success",
                "rotulo": "Sem presença",
                "valor": context["totais"]["sem_presenca"],
                "descricao": "Funcionários sem catraca.",
            },
            "fora_cronograma": {
                "nivel": "warning",
                "rotulo": "Fora cronograma",
                "valor": validacoes.filter(tipo="atividade_fora_cronograma").count(),
                "descricao": "Pendências automáticas de atividades.",
            },
            "alertas": {
                "nivel": "warning" if context["totais"]["alertas"] else "success",
                "rotulo": "Alertas",
                "valor": context["totais"]["alertas"],
                "descricao": "ValidaçÃµes monitoradas.",
            },
        }

        return context




def _rdc_consolidado_summary_rows(ctx):
    labels = {
        "rdcs": "RDCs",
        "atividades": "Atividades",
        "funcionarios": "Funcionários",
        "apontamentos": "Apontamentos",
        "validacoes": "ValidaçÃµes",
        "hh_total": "HH Total",
        "bloqueios": "Bloqueios",
        "alertas": "Alertas",
        "fechados": "Fechados",
        "sem_presenca": "Sem presença na catraca",
    }
    totais = ctx.get("totais") or {}
    return [(labels.get(k, str(k).replace("_", " ").title()), v) for k, v in totais.items()]


def _rdc_consolidado_subtitle(request):
    agora = timezone.localtime().strftime("%d/%m/%Y %H:%M")
    q = (request.GET.get("q") or "").strip() or "Todos"
    projeto = (request.GET.get("projeto") or "").strip() or "Todos"
    disciplina = (request.GET.get("disciplina") or "").strip() or "Todas"
    data_ini = (request.GET.get("data_ini") or "").strip() or "-"
    data_fim = (request.GET.get("data_fim") or "").strip() or "-"
    return (
        f"Emitido em {agora} | Busca: {q} | Projeto: {projeto} | "
        f"Disciplina: {disciplina} | Período: {data_ini} a {data_fim}"
    )


def _rdc_kpi_detalhes_rows(ctx):
    dashboard = ctx.get("dashboard") or {}
    previsto = ctx.get("previsto_realizado") or {}
    executivos = ctx.get("executivos") or {}

    rows = []
    for item in (dashboard.get("hh_por_disciplina") or [])[:10]:
        rows.append([
            "HH por disciplina",
            item.get("rdc__disciplina__nome") or "-",
            item.get("total") or 0,
        ])

    for item in (dashboard.get("top_funcionarios") or [])[:10]:
        rows.append([
            "Top funcionários HH",
            item.get("nome") or "-",
            item.get("hh_total") or 0,
        ])

    rows.extend([
        ["Aderência escopo (%)", "Planejado x realizado", previsto.get("percentual_escopo") or Decimal("0.00")],
        ["Aderência HH (%)", "Consumo de HH", previsto.get("percentual_hh") or Decimal("0.00")],
    ])

    for chave, item in executivos.items():
        rows.append([
            f"KPI {item.get('rotulo') or chave}",
            item.get("descricao") or "",
            item.get("valor") or 0,
        ])
    return rows


def _rdc_consolidado_extra_sheets(ctx):
    return [
        {
            "name": "RDCs",
            "title": "Base de RDCs",
            "headers": ["RDC", "Data", "Projeto", "Disciplina", "Área", "Turno", "Status", "Fechado em"],
            "rows": [
                [
                    r.id,
                    r.data,
                    r.projeto.codigo,
                    r.disciplina.nome,
                    r.area_local.descricao,
                    getattr(r, "get_turno_display", lambda: r.turno)(),
                    getattr(r, "get_status_display", lambda: r.status)(),
                    getattr(r, "fechado_em", None),
                ]
                for r in ctx.get("rdcs", [])
            ],
        },
        {
            "name": "Funcionários",
            "title": "Base de Funcionários",
            "headers": [
                "RDC", "Data", "Projeto", "Nome", "Matrícula", "Função", "Equipe",
                "Hora normal", "Hora extra", "HH total", "Presença catraca", "Elegível", "Bloqueio"
            ],
            "rows": [
                [
                    f.rdc.id,
                    f.rdc.data,
                    f.rdc.projeto.codigo,
                    f.nome,
                    f.matricula,
                    getattr(f.funcao, "nome", ""),
                    getattr(f.equipe, "nome", ""),
                    f.hora_normal,
                    f.hora_extra,
                    f.hh_total,
                    f.presente_catraca,
                    f.elegivel,
                    f.motivo_bloqueio,
                ]
                for f in ctx.get("funcionarios", [])
            ],
        },
        {
            "name": "ValidaçÃµes",
            "title": "Base de ValidaçÃµes",
            "headers": ["RDC", "Data", "Projeto", "Tipo", "Status", "Mensagem", "Referência", "Criado em"],
            "rows": [
                [
                    v.rdc.id,
                    v.rdc.data,
                    v.rdc.projeto.codigo,
                    v.get_tipo_display(),
                    v.get_status_display(),
                    v.mensagem,
                    v.referencia,
                    v.created_at,
                ]
                for v in ctx.get("validacoes", [])
            ],
        },
    ]


class RDCConsolidadoExportView(AuthenticatedTemplateMixin, View):
    def get(self, request, tipo):
        view = RDCConsolidadoView()
        view.request = request
        ctx = view.get_context_data()

        if tipo == "atividades":
            rows = [
                [
                    a.rdc.id,
                    a.rdc.data,
                    a.rdc.projeto.codigo,
                    a.rdc.disciplina.nome,
                    a.rdc.area_local.descricao,
                    a.codigo_atividade,
                    a.descr_atividade,
                    a.codigo_subatividade,
                    a.descr_subatividade,
                    a.qtd_escopo,
                    a.qtd_executada,
                    getattr(a, "get_origem_display", lambda: a.origem)(),
                    "Sim" if a.ativa_no_dia else "Não",
                ]
                for a in ctx["atividades"]
            ]
            return _make_csv_response(
                "rdc_atividades_consolidadas.csv",
                [
                    "RDC",
                    "Data",
                    "Projeto",
                    "Disciplina",
                    "Área",
                    "Código",
                    "Descrição",
                    "Subcódigo",
                    "Subdescrição",
                    "Escopo",
                    "Executada",
                    "Origem",
                    "Ativa",
                ],
                rows,
            )

        if tipo == "funcionarios":
            rows = [
                [
                    f.rdc.id,
                    f.rdc.data,
                    f.rdc.projeto.codigo,
                    f.nome,
                    f.matricula,
                    getattr(f.funcao, "nome", ""),
                    getattr(f.equipe, "nome", ""),
                    f.hora_normal,
                    f.hora_extra,
                    f.hh_total,
                    "Sim" if f.presente_catraca else "Não",
                    "Sim" if f.elegivel else "Não",
                    f.motivo_bloqueio,
                ]
                for f in ctx["funcionarios"]
            ]
            return _make_csv_response(
                "rdc_funcionarios_consolidados.csv",
                [
                    "RDC",
                    "Data",
                    "Projeto",
                    "Nome",
                    "Matrícula",
                    "Função",
                    "Equipe",
                    "Hora normal",
                    "Hora extra",
                    "HH total",
                    "Catraca",
                    "Elegível",
                    "Bloqueio",
                ],
                rows,
            )

        if tipo == "apontamentos":
            rows = [
                [
                    a.rdc.id,
                    a.rdc.data,
                    a.rdc.projeto.codigo,
                    getattr(a.rdc_funcionario, "nome", ""),
                    getattr(a.rdc_funcionario, "matricula", ""),
                    getattr(a.rdc_atividade, "codigo_atividade", ""),
                    getattr(a.rdc_atividade, "descr_atividade", ""),
                    a.horas,
                    a.observacao,
                ]
                for a in ctx["apontamentos"]
            ]
            return _make_csv_response(
                "rdc_apontamentos_consolidados.csv",
                [
                    "RDC",
                    "Data",
                    "Projeto",
                    "Funcionário",
                    "Matrícula",
                    "Cód atividade",
                    "Atividade",
                    "Horas",
                    "ObservAção",
                ],
                rows,
            )

        if tipo == "validacoes":
            rows = [
                [
                    v.rdc.id,
                    v.rdc.data,
                    v.rdc.projeto.codigo,
                    v.get_tipo_display(),
                    v.get_status_display(),
                    v.mensagem,
                    v.referencia,
                    v.created_at,
                ]
                for v in ctx["validacoes"]
            ]
            return _make_csv_response(
                "rdc_validacoes_consolidadas.csv",
                [
                    "RDC",
                    "Data",
                    "Projeto",
                    "Tipo",
                    "Status",
                    "Mensagem",
                    "Referência",
                    "Criado em",
                ],
                rows,
            )

        if tipo == "resumo":
            rows = [
                [
                    r.id,
                    r.data,
                    r.projeto.codigo,
                    r.disciplina.nome,
                    r.area_local.descricao,
                    getattr(r, "get_turno_display", lambda: r.turno)(),
                    getattr(r, "get_status_display", lambda: r.status)(),
                    getattr(r, "fechado_em", None),
                ]
                for r in ctx["rdcs"]
            ]
            return _make_csv_response(
                "rdc_resumo_consolidado.csv",
                ["RDC", "Data", "Projeto", "Disciplina", "Área", "Turno", "Status", "Fechado em"],
                rows,
            )

        if tipo == "excel":
            resumo_rows = _rdc_consolidado_summary_rows(ctx)
            detalhe_rows = _rdc_kpi_detalhes_rows(ctx)
            return _make_xlsx_response(
                "rdc_consolidado.xlsx",
                ["Indicador", "Referência", "Valor"],
                detalhe_rows,
                sheet_name="Dashboard",
                title="RDC Consolidado - Dashboard Gerencial",
                subtitle=_rdc_consolidado_subtitle(request),
                summary_rows=resumo_rows,
                extra_sheets=_rdc_consolidado_extra_sheets(ctx),
            )

        if tipo == "pdf":
            resumo_rows = _rdc_consolidado_summary_rows(ctx)
            detalhe_rows = _rdc_kpi_detalhes_rows(ctx)
            return _make_pdf_response(
                "rdc_consolidado.pdf",
                "RDC Consolidado - Relatório Gerencial",
                ["Indicador", "Referência", "Valor"],
                detalhe_rows,
                subtitle=_rdc_consolidado_subtitle(request),
                summary_rows=resumo_rows,
            )

        raise Http404("Tipo de exportAção não suportado.")


class RDOView(AuthenticatedTemplateMixin, TemplateView):
    template_name = "rdc/rdo.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        scope = _resolve_rdo_scope(self.request)

        from cadastros.models import Projeto, Disciplina

        data_ref = self.request.GET.get("data", "").strip() or timezone.localdate().isoformat()

        context["projetos"] = Projeto.objects.filter(ativo=True).order_by("codigo")
        context["disciplinas"] = Disciplina.objects.filter(ativo=True).order_by("nome")
        context["filter_values"] = {
            "data": data_ref,
            "projeto": self.request.GET.get("projeto", ""),
            "disciplina": self.request.GET.get("disciplina", ""),
            "q": self.request.GET.get("q", ""),
        }
        context["ids_selecionados"] = scope["ids_selecionados"]
        context["rdcs_disponiveis"] = scope["rdcs_filtrados"]
        context["resumo_disciplina"] = scope["resumo_disciplina"]
        context["observacoes_consolidadas"] = scope["observacoes_consolidadas"]
        context["totais"] = scope["totais"]
        return context




def _rdo_summary_rows(scope):
    totais = scope.get("totais") or {}
    labels = [
        ("rdcs", "RDCs"),
        ("atividades", "Atividades"),
        ("funcionarios", "Funcionários"),
        ("apontamentos", "Apontamentos"),
        ("hh_total", "HH Total"),
        ("alertas", "Alertas"),
        ("bloqueios", "Bloqueios"),
    ]
    return [(rotulo, totais.get(chave, 0)) for chave, rotulo in labels]


def _rdo_subtitle(request, scope):
    agora = timezone.localtime().strftime("%d/%m/%Y %H:%M")
    data_ref = (request.GET.get("data") or "").strip() or "Todas"
    projeto = (request.GET.get("projeto") or "").strip() or "Todos"
    disciplina = (request.GET.get("disciplina") or "").strip() or "Todas"
    busca = (request.GET.get("q") or "").strip() or "Todos"
    selecionados = len(scope.get("ids_selecionados") or [])
    return (
        f"Emitido em {agora} | Data: {data_ref} | Projeto: {projeto} | "
        f"Disciplina: {disciplina} | Busca: {busca} | RDCs selecionados: {selecionados}"
    )


def _rdo_detail_rows(scope):
    rows = []
    for item in (scope.get("resumo_disciplina") or [])[:12]:
        rows.append([
            "Resumo por disciplina",
            item.get("rdc__disciplina__nome") or "-",
            f"Ativ.: {item.get('atividades') or 0} | Escopo: {item.get('escopo') or 0} | Executado: {item.get('executado') or 0}",
        ])

    rdcs = list(scope.get("rdcs_selecionados") or [])[:12]
    atividade_map = {
        item["rdc_id"]: item
        for item in scope["atividades"]
        .values("rdc_id")
        .annotate(
            atividades=Count("id"),
            escopo=Sum("qtd_escopo"),
            executado=Sum("qtd_executada"),
        )
    }
    funcionarios_map = {
        item["rdc_id"]: item
        for item in scope["funcionarios"]
        .values("rdc_id")
        .annotate(
            funcionarios=Count("id"),
            hh_total=Sum("hh_total"),
            presentes=Count("id", filter=Q(presente_catraca=True)),
        )
    }

    for r in rdcs:
        a = atividade_map.get(r.id, {})
        f = funcionarios_map.get(r.id, {})
        rows.append([
            "RDC",
            f"{r.id} - {getattr(r.projeto, 'codigo', '')} - {getattr(r.disciplina, 'nome', '')}",
            (
                f"HH: {f.get('hh_total') or 0} | Presentes: {f.get('presentes') or 0}/{f.get('funcionarios') or 0} | "
                f"Ativ.: {a.get('atividades') or 0} | Exec.: {a.get('executado') or 0}"
            ),
        ])

    if not rows:
        rows.append(["Resumo", "Sem dados no filtro", "0"])

    return rows


def _rdo_extra_sheets(scope):
    headers_base = [
        "RDC",
        "Data",
        "Projeto",
        "Disciplina",
        "Área",
        "Turno",
        "Status",
        "Atividades",
        "Escopo",
        "Executado",
        "Funcionários",
        "Presentes",
        "HH total",
        "ObservaçÃµes",
    ]
    base_rows = _rdo_table_rows(scope)

    disciplina_rows = [
        [
            item.get("rdc__disciplina__nome") or "-",
            item.get("atividades") or 0,
            item.get("escopo") or 0,
            item.get("executado") or 0,
        ]
        for item in (scope.get("resumo_disciplina") or [])
    ]

    observacoes_rows = [
        [idx + 1, obs]
        for idx, obs in enumerate(scope.get("observacoes_consolidadas") or [])
    ]

    apontamentos_rows = [
        [
            a.rdc.id,
            a.rdc.data,
            a.rdc.projeto.codigo,
            getattr(a.rdc_funcionario, "nome", ""),
            getattr(a.rdc_atividade, "codigo_atividade", ""),
            getattr(a.rdc_atividade, "descr_atividade", ""),
            a.horas,
            a.observacao,
        ]
        for a in scope.get("apontamentos", [])
    ]

    return [
        {
            "name": "Base RDO",
            "title": "Base do RDO",
            "headers": headers_base,
            "rows": base_rows,
        },
        {
            "name": "Disciplinas",
            "title": "Resumo por Disciplina",
            "headers": ["Disciplina", "Atividades", "Escopo", "Executado"],
            "rows": disciplina_rows,
        },
        {
            "name": "Apontamentos",
            "title": "Base de Apontamentos",
            "headers": ["RDC", "Data", "Projeto", "Funcionário", "Cód. atividade", "Atividade", "Horas", "ObservAção"],
            "rows": apontamentos_rows,
        },
        {
            "name": "ObservaçÃµes",
            "title": "ObservaçÃµes Consolidadas",
            "headers": ["#", "ObservAção"],
            "rows": observacoes_rows,
        },
    ]


class RDOExportView(AuthenticatedTemplateMixin, View):
    def get(self, request, tipo):
        scope = _resolve_rdo_scope(request)
        headers = [
            "RDC",
            "Data",
            "Projeto",
            "Disciplina",
            "Área",
            "Turno",
            "Status",
            "Atividades",
            "Escopo",
            "Executado",
            "Funcionários",
            "Presentes",
            "HH total",
            "ObservaçÃµes",
        ]
        rows = _rdo_table_rows(scope)

        if tipo == "csv":
            return _make_csv_response("rdo.csv", headers, rows)

        if tipo == "excel":
            return _make_xlsx_response(
                "rdo.xlsx",
                ["Indicador", "Referência", "Valor"],
                _rdo_detail_rows(scope),
                sheet_name="Dashboard RDO",
                title="RDO - Dashboard Gerencial",
                subtitle=_rdo_subtitle(request, scope),
                summary_rows=_rdo_summary_rows(scope),
                extra_sheets=_rdo_extra_sheets(scope),
            )

        if tipo == "pdf":
            return _make_pdf_response(
                "rdo.pdf",
                "RDO - Relatório Diário de Obra",
                ["Indicador", "Referência", "Valor"],
                _rdo_detail_rows(scope),
                subtitle=_rdo_subtitle(request, scope),
                summary_rows=_rdo_summary_rows(scope),
            )

        raise Http404("Tipo de exportAção de RDO não suportado.")


class RDCUpdateView(AuthenticatedTemplateMixin, RoleRequiredMixin, UpdateView):
    allowed_roles = ["admin", "supervisor"]
    model = RDC
    form_class = RDCForm
    template_name = "rdc/editar_rdc.html"
    context_object_name = "rdc"

    def form_valid(self, form):
        messages.success(self.request, "RDC atualizado com sucesso.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse("rdc-detail", kwargs={"pk": self.object.pk})


class RDCDeleteView(AuthenticatedTemplateMixin, RoleRequiredMixin, DeleteView):
    allowed_roles = ["admin"]
    model = RDC
    template_name = "rdc/excluir_rdc.html"
    context_object_name = "rdc"

    def form_valid(self, form):
        messages.success(self.request, "RDC excluído com sucesso.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse("rdc-list")


class RDCValidacoesView(AuthenticatedTemplateMixin, TemplateView):
    template_name = "rdc/validacoes.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        rdc = RDC.objects.prefetch_related("validacoes").get(pk=self.kwargs["pk"])
        context["rdc"] = rdc
        context["validacoes"] = rdc.validacoes.all()
        return context


class RDCExportarModeloView(AuthenticatedTemplateMixin, RoleRequiredMixin, View):
    allowed_roles = ["admin", "supervisor"]
    def get(self, request, pk):
        rdc = get_object_or_404(RDC, pk=pk)
        try:
            arquivo = exportar_rdc_para_modelo_excel(rdc)
        except FileNotFoundError as exc:
            raise Http404(str(exc))
        return FileResponse(open(arquivo, "rb"), as_attachment=True, filename=arquivo.name)


class RDCNestedBaseMixin(AuthenticatedTemplateMixin):
    parent_context_name = "rdc"
    parent_pk_url_kwarg = "pk"
    anchor = ""

    def dispatch(self, request, *args, **kwargs):
        self.rdc = get_object_or_404(RDC, pk=kwargs[self.parent_pk_url_kwarg])
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context[self.parent_context_name] = self.rdc
        context["object_label"] = getattr(self, "object_label", "Item")
        context["cancel_url"] = (
            f"{reverse('rdc-detail', kwargs={'pk': self.rdc.pk})}{self.anchor}"
        )
        return context

    def get_success_url(self):
        return f"{reverse('rdc-detail', kwargs={'pk': self.rdc.pk})}{self.anchor}"


class RDCNestedCreateView(RDCNestedBaseMixin, CreateView):
    template_name = "rdc/item_form.html"

    def form_valid(self, form):
        form.instance.rdc = self.rdc
        response = super().form_valid(form)
        _atualizar_validacoes_automaticas(self.rdc)
        messages.success(self.request, f"{self.object_label} incluído(a) com sucesso.")
        return response


class RDCNestedUpdateView(RDCNestedBaseMixin, UpdateView):
    template_name = "rdc/item_form.html"

    def get_queryset(self):
        return self.model.objects.filter(rdc=self.rdc)

    def form_valid(self, form):
        response = super().form_valid(form)
        _atualizar_validacoes_automaticas(self.rdc)
        messages.success(self.request, f"{self.object_label} atualizado(a) com sucesso.")
        return response


RDCInlineUpdateView = RDCNestedUpdateView


class RDCNestedDeleteView(RDCNestedBaseMixin, DeleteView):
    template_name = "rdc/item_confirm_delete.html"

    def get_queryset(self):
        return self.model.objects.filter(rdc=self.rdc)

    def form_valid(self, form):
        response = super().form_valid(form)
        _atualizar_validacoes_automaticas(self.rdc)
        messages.success(self.request, f"{self.object_label} excluído(a) com sucesso.")
        return response


class RDCAtividadeCreateView(RDCNestedCreateView):
    model = RDCAtividade
    form_class = RDCAtividadeForm
    object_label = "Atividade"
    anchor = "#atividades"

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["rdc"] = self.rdc
        return kwargs


class RDCAtividadeUpdateView(RDCNestedUpdateView):
    model = RDCAtividade
    form_class = RDCAtividadeForm
    object_label = "Atividade"
    anchor = "#atividades"
    pk_url_kwarg = "pk2"

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["rdc"] = self.rdc
        return kwargs


class RDCAtividadeDeleteView(RDCNestedDeleteView):
    model = RDCAtividade
    object_label = "Atividade"
    anchor = "#atividades"
    pk_url_kwarg = "pk2"


class RDCFuncionarioCreateView(RDCNestedCreateView):
    model = RDCFuncionario
    form_class = RDCFuncionarioForm
    object_label = "Funcionário"
    anchor = "#funcionarios"

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["rdc"] = self.rdc
        return kwargs


class RDCFuncionarioUpdateView(RDCNestedUpdateView):
    model = RDCFuncionario
    form_class = RDCFuncionarioForm
    object_label = "Funcionário"
    anchor = "#funcionarios"
    pk_url_kwarg = "pk2"

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["rdc"] = self.rdc
        return kwargs


class RDCFuncionarioDeleteView(RDCNestedDeleteView):
    model = RDCFuncionario
    object_label = "Funcionário"
    anchor = "#funcionarios"
    pk_url_kwarg = "pk2"


class RDCApontamentoCreateView(RDCNestedCreateView):
    model = RDCApontamento
    form_class = RDCApontamentoForm
    object_label = "Apontamento"
    anchor = "#apontamentos"

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["rdc"] = self.rdc
        return kwargs


class RDCApontamentoUpdateView(RDCNestedUpdateView):
    model = RDCApontamento
    form_class = RDCApontamentoForm
    object_label = "Apontamento"
    anchor = "#apontamentos"
    pk_url_kwarg = "pk2"

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["rdc"] = self.rdc
        return kwargs


class RDCApontamentoDeleteView(RDCNestedDeleteView):
    model = RDCApontamento
    object_label = "Apontamento"
    anchor = "#apontamentos"
    pk_url_kwarg = "pk2"


class RDCValidacaoCreateView(RDCNestedCreateView):
    model = RDCValidacao
    form_class = RDCValidacaoForm
    object_label = "ValidAção"
    anchor = "#validacoes"


class RDCValidacaoUpdateView(RDCNestedUpdateView):
    model = RDCValidacao
    form_class = RDCValidacaoForm
    object_label = "ValidAção"
    anchor = "#validacoes"
    pk_url_kwarg = "pk2"


class RDCValidacaoDeleteView(RDCNestedDeleteView):
    model = RDCValidacao
    object_label = "ValidAção"
    anchor = "#validacoes"
    pk_url_kwarg = "pk2"


class RDCRevalidarView(AuthenticatedTemplateMixin, View):
    def post(self, request, pk):
        rdc = get_object_or_404(RDC, pk=pk)
        _atualizar_validacoes_automaticas(rdc)
        messages.success(request, "RDC revalidado com sucesso.")
        return redirect(f"{reverse('rdc-detail', kwargs={'pk': rdc.pk})}#validacoes")


class RDCAtividadeLoteView(AuthenticatedTemplateMixin, View):
    def post(self, request, pk):
        rdc = get_object_or_404(RDC, pk=pk)
        ids = request.POST.getlist("ids")
        acao = request.POST.get("acao")
        qs = rdc.atividades.filter(pk__in=ids)
        total = qs.count()

        if acao == "ativar":
            qs.update(ativa_no_dia=True)
            messages.success(request, f"{total} atividade(s) ativada(s).")
        elif acao == "inativar":
            qs.update(ativa_no_dia=False)
            messages.success(request, f"{total} atividade(s) inativada(s).")
        elif acao == "excluir":
            qs.delete()
            messages.success(request, f"{total} atividade(s) excluída(s).")
        else:
            messages.warning(request, "Selecione uma Ação válida para atividades.")

        _atualizar_validacoes_automaticas(rdc)
        return redirect(f"{reverse('rdc-detail', kwargs={'pk': rdc.pk})}#atividades")


class RDCFuncionarioLoteView(AuthenticatedTemplateMixin, View):
    def post(self, request, pk):
        rdc = get_object_or_404(RDC, pk=pk)
        ids = request.POST.getlist("ids")
        acao = request.POST.get("acao")
        qs = rdc.funcionarios.filter(pk__in=ids)
        total = qs.count()

        if acao == "confirmar_supervisor":
            qs.update(confirmado_supervisor=True)
            messages.success(request, f"{total} funcionário(s) confirmado(s) pelo supervisor.")
        elif acao == "liberar":
            qs.update(elegivel=True, motivo_bloqueio="")
            messages.success(request, f"{total} funcionário(s) liberado(s).")
        elif acao == "bloquear_sem_catraca":
            total = qs.filter(presente_catraca=False).update(
                elegivel=False,
                motivo_bloqueio="Sem catraca no dia.",
            )
            messages.success(
                request,
                f"{total} funcionário(s) bloqueado(s) por falta de catraca.",
            )
        elif acao == "excluir":
            qs.delete()
            messages.success(request, f"{total} funcionário(s) excluído(s).")
        else:
            messages.warning(request, "Selecione uma Ação válida para funcionários.")

        _atualizar_validacoes_automaticas(rdc)
        return redirect(f"{reverse('rdc-detail', kwargs={'pk': rdc.pk})}#funcionarios")


class RDCApontamentoLoteView(AuthenticatedTemplateMixin, View):
    def post(self, request, pk):
        rdc = get_object_or_404(RDC, pk=pk)
        ids = request.POST.getlist("ids")
        acao = request.POST.get("acao")
        qs = rdc.apontamentos.filter(pk__in=ids)
        total = qs.count()

        if acao == "excluir":
            qs.delete()
            messages.success(request, f"{total} apontamento(s) excluído(s).")
        else:
            messages.warning(request, "Selecione uma Ação válida para apontamentos.")

        _atualizar_validacoes_automaticas(rdc)
        return redirect(f"{reverse('rdc-detail', kwargs={'pk': rdc.pk})}#apontamentos")


class RDCValidacaoLoteView(AuthenticatedTemplateMixin, View):
    def post(self, request, pk):
        rdc = get_object_or_404(RDC, pk=pk)
        ids = request.POST.getlist("ids")
        acao = request.POST.get("acao")
        qs = rdc.validacoes.filter(pk__in=ids)
        total = qs.count()

        if acao == "excluir":
            qs.delete()
            messages.success(request, f"{total} validAção(Ãµes) excluída(s).")
        else:
            messages.warning(request, "Selecione uma Ação válida para validaçÃµes.")

        _atualizar_validacoes_automaticas(rdc)
        return redirect(f"{reverse('rdc-detail', kwargs={'pk': rdc.pk})}#validacoes")


class RDCAtividadeBuscaView(AuthenticatedTemplateMixin, View):
    def get(self, request, pk):
        from planejamento.models import AtividadeCronograma

        rdc = get_object_or_404(RDC, pk=pk)
        q = (request.GET.get("q") or "").strip()

        qs = AtividadeCronograma.objects.filter(
            projeto=rdc.projeto,
            area_local=rdc.area_local,
            disciplina=rdc.disciplina,
            data_inicio__lte=rdc.data,
            data_fim__gte=rdc.data,
        )
        if q:
            qs = qs.filter(
                Q(codigo_atividade__icontains=q)
                | Q(descr_atividade__icontains=q)
                | Q(codigo_subatividade__icontains=q)
                | Q(descr_subatividade__icontains=q)
            )

        qs = qs.order_by("codigo_atividade", "descr_atividade")[:30]

        return JsonResponse(
            {
                "results": [
                    {
                        "id": i.id,
                        "texto": f"{i.codigo_atividade or ''} - {i.descr_atividade or ''}".strip(" -"),
                        "codigo_atividade": i.codigo_atividade or "",
                        "descr_atividade": i.descr_atividade or "",
                        "codigo_subatividade": i.codigo_subatividade or "",
                        "descr_subatividade": i.descr_subatividade or "",
                        "qtd_escopo": str(i.qtd_escopo or ""),
                        "badge": "Cronograma",
                        "meta": " / ".join(
                            [x for x in [i.codigo_subatividade or "", i.descr_subatividade or ""] if x]
                        ),
                    }
                    for i in qs
                ]
            }
        )



def _resolve_funcionario_autofill_payload(rdc, funcionario, equipe_id=None, equipe_nome=""):
    alertas = []
    funcao_nome = getattr(getattr(funcionario, "funcao", None), "nome", "") or ""
    presente = False
    elegivel = True
    motivo_bloqueio = ""

    try:
        from acesso.models import RegistroCatraca

        if funcionario.matricula:
            presente = RegistroCatraca.objects.filter(
                matricula=funcionario.matricula,
                data=rdc.data,
                presente=True,
            ).exists()
        if not presente:
            elegivel = False
            alertas.append("Sem catraca no dia.")
    except Exception:
        alertas.append("Não foi possível validar a catraca do dia.")

    if not equipe_id:
        alertas.append("Equipe/alocAção não encontrada automaticamente para o contexto do RDC.")

    motivo_bloqueio = " ; ".join(alertas)
    hora_normal = Decimal("8.00")
    hora_extra = Decimal("0.00")
    hh_total = hora_normal + hora_extra

    return {
        "id": funcionario.id,
        "texto": f"{funcionario.matricula or ''} - {funcionario.nome or ''}".strip(" -"),
        "matricula": funcionario.matricula or "",
        "nome": funcionario.nome or "",
        "funcao_id": getattr(funcionario, "funcao_id", None),
        "funcao_nome": funcao_nome,
        "equipe_id": equipe_id,
        "equipe_nome": equipe_nome or "",
        "presente_catraca": presente,
        "hora_normal": f"{hora_normal:.2f}",
        "hora_extra": f"{hora_extra:.2f}",
        "hh_total": f"{hh_total:.2f}",
        "elegivel": bool(elegivel),
        "motivo_bloqueio": motivo_bloqueio,
        "badge": "Pronto" if presente and equipe_id else "Revisar",
        "meta": " / ".join([x for x in [funcao_nome, equipe_nome] if x]),
        "alertas": alertas,
        "sem_catraca": not presente,
        "sem_alocacao": not bool(equipe_id),
        "diagnostico": {
            "cadastro_ok": bool(funcionario.nome and funcionario.matricula),
            "funcao_ok": bool(funcao_nome),
            "alocacao_ok": bool(equipe_id),
            "catraca_ok": bool(presente),
        },
    }


class RDCFuncionarioBuscaView(AuthenticatedTemplateMixin, View):
    def get(self, request, pk):
        from cadastros.models import Funcionario

        rdc = get_object_or_404(RDC, pk=pk)
        q = (request.GET.get("q") or "").strip()

        qs = Funcionario.objects.select_related("funcao").filter(ativo=True)

        try:
            from alocacao.models import FuncionarioProjeto

            alocados = (
                FuncionarioProjeto.objects.select_related("equipe")
                .filter(
                    projeto=rdc.projeto,
                    disciplina=rdc.disciplina,
                    ativo=True,
                    data_inicio__lte=rdc.data,
                )
                .filter(Q(data_fim__isnull=True) | Q(data_fim__gte=rdc.data))
            )
            ids_alocados = list(alocados.values_list("funcionario_id", flat=True).distinct())
            mapa_equipes = {a.funcionario_id: getattr(a.equipe, "id", None) for a in alocados}
            mapa_equipes_nome = {a.funcionario_id: getattr(a.equipe, "nome", "") for a in alocados}
        except Exception:
            ids_alocados = []
            mapa_equipes = {}
            mapa_equipes_nome = {}

        if q:
            qs = qs.filter(Q(nome__icontains=q) | Q(matricula__icontains=q))
        elif ids_alocados:
            # sem texto digitado, prioriza os alocados para não trazer universo muito grande
            qs = qs.filter(id__in=ids_alocados)

        qs = qs.order_by("nome")[:30]

        results = []
        for funcionario in qs:
            payload = _resolve_funcionario_autofill_payload(
                rdc,
                funcionario,
                equipe_id=mapa_equipes.get(funcionario.id),
                equipe_nome=mapa_equipes_nome.get(funcionario.id, ""),
            )
            if funcionario.id in ids_alocados:
                payload["badge"] = "Alocado" if payload["presente_catraca"] else "Alocado sem catraca"
            results.append(payload)

        return JsonResponse({"results": results})



class RDCImportarAtividadesCronogramaView(AuthenticatedTemplateMixin, View):
    def post(self, request, pk):
        from planejamento.models import AtividadeCronograma

        rdc = get_object_or_404(RDC, pk=pk)
        qs = AtividadeCronograma.objects.filter(
            projeto=rdc.projeto,
            area_local=rdc.area_local,
            disciplina=rdc.disciplina,
            data_inicio__lte=rdc.data,
            data_fim__gte=rdc.data,
        ).order_by("codigo_atividade", "descr_atividade")

        existentes = set(
            rdc.atividades.values_list(
                "codigo_atividade",
                "descr_atividade",
                "codigo_subatividade",
                "descr_subatividade",
            )
        )
        criadas = 0

        for atividade in qs:
            chave = (
                atividade.codigo_atividade,
                atividade.descr_atividade,
                atividade.codigo_subatividade,
                atividade.descr_subatividade,
            )
            if chave in existentes:
                continue

            RDCAtividade.objects.create(
                rdc=rdc,
                atividade_cronograma=atividade,
                codigo_atividade=atividade.codigo_atividade,
                descr_atividade=atividade.descr_atividade,
                codigo_subatividade=atividade.codigo_subatividade,
                descr_subatividade=atividade.descr_subatividade,
                qtd_escopo=getattr(atividade, "qtd_escopo", None),
                origem=getattr(atividade, "origem", None)
                or getattr(RDCAtividade._meta.get_field("origem"), "default", ""),
                obrigatoria=True,
                ativa_no_dia=True,
            )
            criadas += 1

        _atualizar_validacoes_automaticas(rdc)
        messages.success(request, f"{criadas} atividade(s) importada(s) do cronograma.")
        return redirect(f"{reverse('rdc-detail', kwargs={'pk': rdc.pk})}#atividades")


class RDCImportarFuncionariosAlocacaoView(AuthenticatedTemplateMixin, View):
    def post(self, request, pk):
        from alocacao.models import FuncionarioProjeto
        from acesso.models import RegistroCatraca

        rdc = get_object_or_404(RDC, pk=pk)
        alocacoes = (
            FuncionarioProjeto.objects.select_related("funcionario", "funcionario__funcao", "equipe")
            .filter(
                projeto=rdc.projeto,
                disciplina=rdc.disciplina,
                ativo=True,
                data_inicio__lte=rdc.data,
            )
            .filter(Q(data_fim__isnull=True) | Q(data_fim__gte=rdc.data))
            .order_by("funcionario__nome")
        )
        existentes = set(rdc.funcionarios.values_list("funcionario_id", flat=True))
        presentes = set(
            RegistroCatraca.objects.filter(data=rdc.data, presente=True).values_list(
                "matricula",
                flat=True,
            )
        )
        criados = 0

        for aloc in alocacoes:
            funcionario = aloc.funcionario
            if not funcionario or funcionario.id in existentes:
                continue

            presente = (funcionario.matricula or "") in presentes
            RDCFuncionario.objects.create(
                rdc=rdc,
                funcionario=funcionario,
                equipe=aloc.equipe,
                funcao=getattr(funcionario, "funcao", None),
                matricula=funcionario.matricula,
                nome=funcionario.nome,
                hora_normal=Decimal("8.00"),
                hora_extra=Decimal("0.00"),
                hh_total=Decimal("8.00"),
                presente_catraca=presente,
                elegivel=presente,
                motivo_bloqueio="" if presente else "Sem catraca no dia.",
            )
            criados += 1

        _atualizar_validacoes_automaticas(rdc)
        messages.success(request, f"{criados} funcionário(s) importado(s) da alocAção.")
        return redirect(f"{reverse('rdc-detail', kwargs={'pk': rdc.pk})}#funcionarios")


class RDCExportView(RDCExportarModeloView):
    pass



class RDCDashboardHomeView(AuthenticatedTemplateMixin, TemplateView):
    template_name = "rdc/dashboard.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(build_rdc_dashboard_home_context())
        return context



class RDCWorkflowView(AuthenticatedTemplateMixin, RoleRequiredMixin, View):
    allowed_roles = ["admin", "supervisor"]
    def post(self, request, *args, **kwargs):
        rdc = get_object_or_404(RDC, pk=kwargs["pk"])
        acao = (request.POST.get("acao") or "").strip()
        observacao = (request.POST.get("observacao") or "").strip()

        resultado = process_rdc_workflow_action(
            rdc,
            action=acao,
            user=request.user,
            observacao=observacao,
        )

        if resultado["ok"]:
            messages.success(request, resultado["message"])
        else:
            messages.warning(request, resultado["message"])

        if observacao and not resultado["ok"]:
            messages.info(request, f"ObservAção: {observacao}")
        return redirect("rdc-detail", pk=rdc.pk)





























