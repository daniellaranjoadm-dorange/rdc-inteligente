from collections import Counter, defaultdict
from decimal import Decimal
from pathlib import Path
from datetime import datetime

from django.conf import settings
from django.db import transaction
from django.db.models import Q

from acesso.models import RegistroCatraca
from alocacao.models import FuncionarioProjeto, HistogramaPlanejado
from core.choices import (
    OrigemAtividadeChoices,
    StatusRDCChoices,
    StatusValidacaoChoices,
    TipoValidacaoChoices,
    TurnoChoices,
)
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment
from planejamento.models import AtividadeCronograma
from rdc.models import RDC, RDCAtividade, RDCFuncionario, RDCValidacao
from .rdc_excel_layout_service import (
    CIDADE_OBRA,
    HORARIO_INICIO_PADRAO,
    HORARIO_INTERVALO_INICIO_PADRAO,
    HORARIO_INTERVALO_FIM_PADRAO,
    HORARIO_TERMINO_PADRAO,
    SUPERVISOR_PADRAO_NOME,
    EQUIPE_TESTE_PRIORITARIA,
    FILL_INFO,
    FILL_ALERTA,
    FILL_DESTAQUE,
    FILL_CABECALHO,
    FILL_SUPERVISOR,
    FILL_EQUIPE,
    FILL_LIDERANCA,
    FILL_RODAPE,
    FILL_TITULO_FORTE,
    FILL_HEADER_BOX,
    FILL_HEADER_BOX_2,
    _normalizar_texto,
    _resolver_celula_editavel,
    _aplicar_estilo_celula,
    _marcar_celula_climatica,
    inserir_logo,
    ajustar_layout_topo,
)




def buscar_atividades_planejadas(projeto_id, area_local_id, disciplina_id, data, turno):
    turnos_aceitos = [turno, TurnoChoices.INTEGRAL]
    return AtividadeCronograma.objects.filter(
        projeto_id=projeto_id,
        area_local_id=area_local_id,
        disciplina_id=disciplina_id,
        data_inicio__lte=data,
        data_fim__gte=data,
        turno__in=turnos_aceitos,
    )


def buscar_histograma_do_dia(projeto_id, area_local_id, disciplina_id, data, turno):
    turnos_aceitos = [turno, TurnoChoices.INTEGRAL]
    return HistogramaPlanejado.objects.select_related("equipe", "funcao").filter(
        projeto_id=projeto_id,
        area_local_id=area_local_id,
        disciplina_id=disciplina_id,
        data=data,
        turno__in=turnos_aceitos,
    )


def buscar_funcionarios_alocados(projeto_id, disciplina_id, data, equipe_id=None):
    filtros = Q(
        projeto_id=projeto_id,
        disciplina_id=disciplina_id,
        ativo=True,
        data_inicio__lte=data,
    ) & (Q(data_fim__isnull=True) | Q(data_fim__gte=data))

    if equipe_id:
        filtros &= Q(equipe_id=equipe_id)

    return FuncionarioProjeto.objects.select_related(
        "funcionario",
        "funcionario__funcao",
        "equipe",
    ).filter(filtros)


def validar_presenca_catraca(matricula, data):
    return RegistroCatraca.objects.filter(
        matricula=matricula,
        data=data,
        presente=True,
    ).exists()


def registrar_validacao(rdc, tipo, status, mensagem, referencia=""):
    return RDCValidacao.objects.create(
        rdc=rdc,
        tipo=tipo,
        status=status,
        mensagem=mensagem,
        referencia=referencia,
    )


def _resetar_rdc_existente(rdc: RDC) -> RDC:
    rdc.apontamentos.all().delete()
    rdc.funcionarios.all().delete()
    rdc.atividades.all().delete()
    rdc.validacoes.all().delete()
    rdc.status = StatusRDCChoices.PRE_PREENCHIDO
    rdc.save(update_fields=["status", "updated_at"])
    return rdc


def _traduzir_weather_code(weather_code):
    return clima_service._traduzir_weather_code(weather_code)


def _classificar_clima_visual(descricao):
    return clima_service._classificar_clima_visual(descricao)


def buscar_clima_online_rio_grande(data):
    return clima_service.buscar_clima_online_rio_grande(data)


def buscar_supervisor_padrao_ou_equipe(rdc):
    from cadastros.models import Funcionario

    supervisor = Funcionario.objects.filter(nome__iexact=SUPERVISOR_PADRAO_NOME).first()
    if supervisor:
        return supervisor

    funcionario_rdc = (
        rdc.funcionarios.select_related("funcao", "funcionario")
        .filter(funcao__nome__icontains="SUPERVISOR")
        .first()
    )
    if funcionario_rdc and funcionario_rdc.funcionario:
        return funcionario_rdc.funcionario

    alocacao_supervisor = (
        FuncionarioProjeto.objects.select_related("funcionario", "funcionario__funcao", "equipe")
        .filter(projeto=rdc.projeto, disciplina=rdc.disciplina, ativo=True)
        .filter(Q(data_fim__isnull=True) | Q(data_fim__gte=rdc.data))
        .filter(funcionario__funcao__nome__icontains="SUPERVISOR")
        .order_by("funcionario__nome")
        .first()
    )
    if alocacao_supervisor:
        return alocacao_supervisor.funcionario

    return None


def popular_metadados_profissionais_do_rdc(rdc):
    supervisor = buscar_supervisor_padrao_ou_equipe(rdc)
    clima = buscar_clima_online_rio_grande(rdc.data)

    observacoes_linhas = [
        f"Obra: {CIDADE_OBRA}",
        f"Projeto: {rdc.projeto.codigo}",
        f"Área/local: {rdc.area_local.descricao}",
        f"Horário: {HORARIO_INICIO_PADRAO} | Intervalo {HORARIO_INTERVALO_INICIO_PADRAO}-{HORARIO_INTERVALO_FIM_PADRAO} | Término {HORARIO_TERMINO_PADRAO}",
        f"Clima: {clima.get('descricao')}",
    ]

    if clima.get("temperatura_max") is not None and clima.get("temperatura_min") is not None:
        observacoes_linhas.append(
            f"Temperatura: mín {clima['temperatura_min']}°C / máx {clima['temperatura_max']}°C"
        )
    if clima.get("chuva_mm") is not None:
        observacoes_linhas.append(f"PrecipitAção: {clima['chuva_mm']} mm")
    if clima.get("vento_kmh") is not None:
        observacoes_linhas.append(f"Vento: {clima['vento_kmh']} km/h")

    rdc.condicao_area = "LIBERADA"
    rdc.observacoes = " | ".join(observacoes_linhas)

    update_fields = ["condicao_area", "observacoes"]
    if supervisor and hasattr(rdc, "supervisor_id"):
        rdc.supervisor = supervisor
        update_fields.append("supervisor")

    rdc.save(update_fields=update_fields)
    return clima, supervisor


def _peso_funcao(nome_funcao):
    txt = str(nome_funcao or "").upper()
    if "SUPERVISOR" in txt:
        return 1
    if "LÃDER" in txt or "LIDER" in txt:
        return 2
    if "ENCARREGADO" in txt:
        return 3
    if "ENGENHEIRO" in txt:
        return 4
    if "TÃ‰CNICO" in txt or "TECNICO" in txt:
        return 5
    if "MONTADOR" in txt or "SOLDADOR" in txt or "PINTOR" in txt or "ELETRICISTA" in txt:
        return 6
    if "AJUDANTE" in txt or "AUXILIAR" in txt:
        return 7
    return 8


def _funcao_eh_lideranca(nome_funcao):
    txt = str(nome_funcao or "").upper()
    return any(x in txt for x in ["SUPERVISOR", "LÃDER", "LIDER", "ENCARREGADO", "ENGENHEIRO"])


def _alocacoes_priorizadas(alocacoes_queryset):
    alocacoes = list(alocacoes_queryset)

    def chave(aloc):
        equipe_codigo = getattr(aloc.equipe, "codigo", "") or ""
        equipe_prioritaria = 0 if equipe_codigo == EQUIPE_TESTE_PRIORITARIA else 1
        funcao_nome = getattr(getattr(aloc.funcionario, "funcao", None), "nome", "") or ""
        return (equipe_prioritaria, _peso_funcao(funcao_nome), (aloc.funcionario.nome or "").upper())

    return sorted(alocacoes, key=chave)


def popular_funcionarios_rdc_por_alocacao_sem_histograma(rdc, projeto_id, disciplina_id, data):
    alocacoes_base = (
        FuncionarioProjeto.objects.select_related(
            "funcionario",
            "funcionario__funcao",
            "equipe",
        )
        .filter(
            projeto_id=projeto_id,
            disciplina_id=disciplina_id,
            ativo=True,
            data_inicio__lte=data,
        )
        .filter(Q(data_fim__isnull=True) | Q(data_fim__gte=data))
    )

    alocacoes = _alocacoes_priorizadas(alocacoes_base)
    vistos = set()

    for alocacao in alocacoes:
        funcionario = alocacao.funcionario
        if not funcionario or funcionario.id in vistos:
            continue

        presente = validar_presenca_catraca(funcionario.matricula, data)
        RDCFuncionario.objects.create(
            rdc=rdc,
            funcionario=funcionario,
            equipe=alocacao.equipe,
            funcao=funcionario.funcao,
            matricula=funcionario.matricula,
            nome=funcionario.nome,
            hora_normal=Decimal("8.00"),
            hora_extra=Decimal("0.00"),
            hh_total=Decimal("8.00"),
            presente_catraca=presente,
            elegivel=True,
            motivo_bloqueio="" if presente else "Sem catraca no dia - incluído por simulAção de alocAção",
        )
        vistos.add(funcionario.id)

        if not presente:
            registrar_validacao(
                rdc,
                TipoValidacaoChoices.FUNCIONARIO_SEM_CATRACA,
                StatusValidacaoChoices.ALERTA,
                f"{funcionario.nome} sem catraca em {data}, incluído por simulAção de alocAção.",
                referencia=funcionario.matricula,
            )

    if not vistos:
        registrar_validacao(
            rdc,
            TipoValidacaoChoices.FUNCIONARIO_SEM_ALOCACAO,
            StatusValidacaoChoices.ALERTA,
            "Nenhum funcionário alocado foi encontrado para o contexto.",
            referencia=f"{projeto_id}-{disciplina_id}-{data}",
        )

    return len(vistos)


def distribuir_horas_por_atividades(funcionario, atividades):
    total_atividades = len(atividades)
    if total_atividades <= 0:
        return []

    nome_funcao = getattr(getattr(funcionario, "funcao", None), "nome", "")
    lideranca = _funcao_eh_lideranca(nome_funcao)
    letras_disponiveis = ["A", "B", "C", "D", "E", "F"][:total_atividades]

    if total_atividades == 1:
        return [(letras_disponiveis[0], Decimal("8.00"))]

    if total_atividades == 2:
        if lideranca:
            return [(letras_disponiveis[0], Decimal("4.00")), (letras_disponiveis[1], Decimal("4.00"))]
        return [(letras_disponiveis[0], Decimal("5.00")), (letras_disponiveis[1], Decimal("3.00"))]

    if lideranca:
        return [
            (letras_disponiveis[0], Decimal("2.00")),
            (letras_disponiveis[1], Decimal("2.00")),
            (letras_disponiveis[2], Decimal("4.00")),
        ]

    return [
        (letras_disponiveis[0], Decimal("4.00")),
        (letras_disponiveis[1], Decimal("2.00")),
        (letras_disponiveis[2], Decimal("2.00")),
    ]


@transaction.atomic
def montar_rdc_pre_preenchido(projeto_id, area_local_id, disciplina_id, data, turno, user, recriar=True):
    rdc = RDC.objects.filter(
        projeto_id=projeto_id,
        area_local_id=area_local_id,
        disciplina_id=disciplina_id,
        data=data,
        turno=turno,
    ).first()

    if rdc and recriar:
        rdc = _resetar_rdc_existente(rdc)
    elif not rdc:
        rdc = RDC.objects.create(
            projeto_id=projeto_id,
            area_local_id=area_local_id,
            disciplina_id=disciplina_id,
            data=data,
            turno=turno,
            dia_semana=data.strftime("%A"),
            status=StatusRDCChoices.PRE_PREENCHIDO,
            criado_por=user,
        )

    programacoes, _cal = buscar_programacao_semana(
        projeto_id,
        data,
        disciplina_id=disciplina_id,
        area_local_id=area_local_id,
        turno=turno,
    )

    usados = set()
    for prog in programacoes:
        chave = (prog.codigo_atividade or "", prog.codigo_subatividade or "")
        if chave in usados:
            continue

        RDCAtividade.objects.create(
            rdc=rdc,
            atividade_cronograma=prog.atividade_cronograma,
            codigo_atividade=prog.codigo_atividade,
            descr_atividade=prog.descr_atividade,
            codigo_subatividade=prog.codigo_subatividade,
            descr_subatividade=prog.descr_subatividade,
            qtd_escopo=prog.qtd_prevista or Decimal("0.00"),
            origem=OrigemAtividadeChoices.PLANEJAMENTO,
            obrigatoria=True,
            ativa_no_dia=True,
            comentarios=(prog.observacao or ""),
        )
        usados.add(chave)

    atividades = list(buscar_atividades_planejadas(projeto_id, area_local_id, disciplina_id, data, turno))
    for atividade in atividades:
        chave = (atividade.codigo_atividade or "", atividade.codigo_subatividade or "")
        if chave in usados:
            continue

        RDCAtividade.objects.create(
            rdc=rdc,
            atividade_cronograma=atividade,
            codigo_atividade=atividade.codigo_atividade,
            descr_atividade=atividade.descr_atividade,
            codigo_subatividade=atividade.codigo_subatividade,
            descr_subatividade=atividade.descr_subatividade,
            qtd_escopo=atividade.qtd_escopo,
            origem=OrigemAtividadeChoices.CRONOGRAMA,
            obrigatoria=True,
            ativa_no_dia=True,
        )

    if not atividades:
        registrar_validacao(
            rdc,
            TipoValidacaoChoices.ATIVIDADE_FORA_CRONOGRAMA,
            StatusValidacaoChoices.BLOQUEIO,
            "Nenhuma atividade do cronograma encontrada para o contexto informado.",
            referencia=f"{projeto_id}-{area_local_id}-{disciplina_id}-{data}-{turno}",
        )

    histogramas = list(buscar_histograma_do_dia(projeto_id, area_local_id, disciplina_id, data, turno))
    if not histogramas:
        registrar_validacao(
            rdc,
            TipoValidacaoChoices.EQUIPE_FORA_HISTOGRAMA,
            StatusValidacaoChoices.ALERTA,
            "Nenhum histograma planejado encontrado para o contexto informado. O RDC será preenchido por alocAção, com catraca apenas como referência.",
            referencia=f"{projeto_id}-{area_local_id}-{disciplina_id}-{data}-{turno}",
        )
        popular_funcionarios_rdc_por_alocacao_sem_histograma(
            rdc=rdc,
            projeto_id=projeto_id,
            disciplina_id=disciplina_id,
            data=data,
        )
        gerar_apontamentos_base_automaticos(rdc)
        popular_metadados_profissionais_do_rdc(rdc)
        return rdc

    vistos = set()
    for hist in histogramas:
        alocacoes = buscar_funcionarios_alocados(projeto_id, disciplina_id, data, hist.equipe_id)
        total_encontrado = 0

        for alocacao in _alocacoes_priorizadas(alocacoes):
            funcionario = alocacao.funcionario
            if funcionario.id in vistos:
                registrar_validacao(
                    rdc,
                    TipoValidacaoChoices.DUPLICIDADE_APONTAMENTO,
                    StatusValidacaoChoices.BLOQUEIO,
                    f"Funcionário {funcionario.nome} apareceu mais de uma vez no mesmo RDC.",
                    referencia=funcionario.matricula,
                )
                continue

            elegivel = True
            motivo_bloqueio = ""
            presente = validar_presenca_catraca(funcionario.matricula, data)
            total_encontrado += 1

            if not funcionario.ativo:
                elegivel = False
                motivo_bloqueio = "Funcionário inativo."
            elif not alocacao.esta_valido_em(data):
                elegivel = False
                motivo_bloqueio = "Funcionário sem alocAção ativa no contexto."
                registrar_validacao(
                    rdc,
                    TipoValidacaoChoices.FUNCIONARIO_SEM_ALOCACAO,
                    StatusValidacaoChoices.BLOQUEIO,
                    f"{funcionario.nome} sem alocAção ativa para o projeto/disciplina/data.",
                    referencia=funcionario.matricula,
                )
            elif not presente:
                elegivel = False
                motivo_bloqueio = "Sem registro de catraca no dia."
                registrar_validacao(
                    rdc,
                    TipoValidacaoChoices.FUNCIONARIO_SEM_CATRACA,
                    StatusValidacaoChoices.BLOQUEIO,
                    f"{funcionario.nome} não possui presença de catraca em {data}.",
                    referencia=funcionario.matricula,
                )

            if funcionario.funcao_id != hist.funcao_id:
                registrar_validacao(
                    rdc,
                    TipoValidacaoChoices.FUNCAO_DIVERGENTE,
                    StatusValidacaoChoices.ALERTA,
                    f"Função divergente para {funcionario.nome}: cadastrado em {funcionario.funcao} e histograma em {hist.funcao}.",
                    referencia=funcionario.matricula,
                )

            RDCFuncionario.objects.create(
                rdc=rdc,
                funcionario=funcionario,
                equipe=alocacao.equipe,
                funcao=funcionario.funcao,
                matricula=funcionario.matricula,
                nome=funcionario.nome,
                hora_normal=Decimal("8.00") if presente else Decimal("0.00"),
                hora_extra=Decimal("0.00"),
                hh_total=Decimal("8.00") if presente else Decimal("0.00"),
                presente_catraca=presente,
                elegivel=elegivel,
                motivo_bloqueio=motivo_bloqueio,
            )
            vistos.add(funcionario.id)

        if total_encontrado == 0:
            registrar_validacao(
                rdc,
                TipoValidacaoChoices.EQUIPE_FORA_HISTOGRAMA,
                StatusValidacaoChoices.ALERTA,
                f"Nenhum funcionário alocado encontrado para equipe {hist.equipe}.",
                referencia=str(hist.equipe_id),
            )

    gerar_apontamentos_base_automaticos(rdc)
    popular_metadados_profissionais_do_rdc(rdc)
    return rdc

def sugerir_contexto_rdc_por_cronograma(
    projeto_id=None,
    data=None,
    disciplina_id=None,
    area_local_id=None,
    turno=TurnoChoices.INTEGRAL,
):
    qs = AtividadeCronograma.objects.select_related("projeto", "area_local", "disciplina")

    if projeto_id:
        qs = qs.filter(projeto_id=projeto_id)
    if data:
        qs = qs.filter(data_inicio__lte=data, data_fim__gte=data)
    if disciplina_id:
        qs = qs.filter(disciplina_id=disciplina_id)
    if area_local_id:
        qs = qs.filter(area_local_id=area_local_id)

    atividades = list(qs)
    if not atividades:
        raise ValueError("Nenhuma atividade de cronograma encontrada para sugerir um contexto de RDC.")

    if not data:
        contagem_por_data = Counter()
        for atividade in atividades:
            contagem_por_data[atividade.data_inicio] += 1
        data = contagem_por_data.most_common(1)[0][0]
        atividades = [a for a in atividades if a.data_inicio <= data <= a.data_fim]

    if not atividades:
        raise ValueError("Nenhuma atividade de cronograma ativa na data selecionada.")

    if not disciplina_id:
        disciplina_id = Counter(a.disciplina_id for a in atividades).most_common(1)[0][0]
        atividades = [a for a in atividades if a.disciplina_id == disciplina_id]

    if not area_local_id:
        area_local_id = Counter(a.area_local_id for a in atividades).most_common(1)[0][0]
        atividades = [a for a in atividades if a.area_local_id == area_local_id]

    atividade_ref = atividades[0]
    return {
        "projeto": atividade_ref.projeto,
        "data": data,
        "disciplina": atividade_ref.disciplina,
        "area_local": atividade_ref.area_local,
        "turno": turno,
        "total_atividades": len(atividades),
    }




def gerar_apontamentos_base_automaticos(rdc):
    if rdc.apontamentos.exists():
        return 0

    atividades = list(rdc.atividades.filter(ativa_no_dia=True).order_by("codigo_atividade", "codigo_subatividade"))
    funcionarios = [
        item
        for item in rdc.funcionarios.select_related("funcao").all()
        if getattr(item, "pode_apontar", False) and (item.hh_total or Decimal("0.00")) > 0
    ]

    if not atividades or not funcionarios:
        return 0

    mapa_atividades = {chr(65 + idx): atividade for idx, atividade in enumerate(atividades[:6])}
    criados = 0

    for funcionario in funcionarios:
        distribuicao = distribuir_horas_por_atividades(funcionario, list(mapa_atividades.values()))
        for letra, horas in distribuicao:
            atividade = mapa_atividades.get(letra)
            if not atividade or not horas or horas <= 0:
                continue
            _, created = RDCApontamento.objects.get_or_create(
                rdc=rdc,
                rdc_funcionario=funcionario,
                rdc_atividade=atividade,
                defaults={
                    "horas": horas,
                    "observacao": "Gerado automaticamente na montagem assistida.",
                },
            )
            if created:
                criados += 1

    if criados:
        registrar_validacao(
            rdc,
            TipoValidacaoChoices.ATIVIDADE_FORA_CRONOGRAMA,
            StatusValidacaoChoices.INFO,
            f"Montagem assistida gerou {criados} apontamento(s) base para acelerar o preenchimento do RDC.",
            referencia=f"AUTO:APONTAMENTOS_BASE:{rdc.pk}",
        )

    return criados

@transaction.atomic
def montar_rdc_simulado_por_cronograma(
    user,
    projeto_id=None,
    data=None,
    disciplina_id=None,
    area_local_id=None,
    turno=TurnoChoices.INTEGRAL,
):
    contexto = sugerir_contexto_rdc_por_cronograma(
        projeto_id=projeto_id,
        data=data,
        disciplina_id=disciplina_id,
        area_local_id=area_local_id,
        turno=turno,
    )

    rdc = montar_rdc_pre_preenchido(
        projeto_id=contexto["projeto"].id,
        area_local_id=contexto["area_local"].id,
        disciplina_id=contexto["disciplina"].id,
        data=contexto["data"],
        turno=turno,
        user=user,
        recriar=True,
    )

    registrar_validacao(
        rdc,
        TipoValidacaoChoices.ATIVIDADE_FORA_CRONOGRAMA,
        StatusValidacaoChoices.INFO,
        f"RDC simulado montado a partir do cronograma. Atividades ativas no contexto: {contexto['total_atividades']}.",
        referencia=f"{contexto['projeto'].codigo}-{contexto['disciplina'].codigo}-{contexto['area_local'].codigo}",
    )
    return rdc, contexto


def selecionar_atividades_unicas_para_exportacao(rdc, limite=6):
    atividades = list(rdc.atividades.select_related("atividade_cronograma").all().order_by("id"))
    unicas = []
    chaves_vistas = set()

    for atividade in atividades:
        chave = (
            (atividade.codigo_atividade or "").strip().upper(),
            (atividade.descr_atividade or "").strip().upper(),
            (atividade.codigo_subatividade or "").strip().upper(),
            (atividade.descr_subatividade or "").strip().upper(),
        )
        if chave in chaves_vistas:
            continue
        chaves_vistas.add(chave)
        unicas.append(atividade)
        if len(unicas) >= limite:
            break

    return unicas


def buscar_funcionarios_reais_por_catraca_para_rdc(data, limite=8):
    from cadastros.models import Funcionario

    registros = (
        RegistroCatraca.objects.filter(data=data, presente=True)
        .exclude(matricula__isnull=True)
        .exclude(matricula="")
        .order_by("matricula")
    )
    matriculas = list(registros.values_list("matricula", flat=True).distinct())

    funcionarios = (
        Funcionario.objects.select_related("funcao")
        .filter(matricula__in=matriculas, ativo=True)
        .order_by("nome")[:limite]
    )
    return list(funcionarios)




def _localizar_celula_por_texto(ws, texto_exato):
    alvo = _normalizar_texto(texto_exato)
    for row in ws.iter_rows():
        for cell in row:
            if _normalizar_texto(cell.value) == alvo:
                return cell
    return None



    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col)

    return cell






def _equipe_principal_nome(rdc):
    equipe = rdc.funcionarios.exclude(equipe__isnull=True).values_list("equipe__nome", flat=True).first()
    return equipe or "Equipe não identificada"


def _contar_liderancas(rdc):
    total = 0
    for f in rdc.funcionarios.select_related("funcao").all():
        nome_funcao = getattr(f.funcao, "nome", "") or ""
        if _funcao_eh_lideranca(nome_funcao):
            total += 1
    return total


def _hh_total_rdc(rdc):
    total = Decimal("0.00")
    for f in rdc.funcionarios.all():
        total += f.hh_total or Decimal("0.00")
    return total





    try:
        img = XLImage(str(logo_path))
        img.width = 185
        img.height = 78
        ws.add_image(img, "A1")
        print(f"[RDC] Logo inserida com sucesso: {logo_path}")
        return True
    except Exception as exc:
        print(f"[RDC] Falha ao inserir logo '{logo_path}': {exc}")
        return False




def preencher_quadro_climatico(ws, clima_info):
    visual = clima_info.get("visual") or _classificar_clima_visual(clima_info.get("descricao"))
    coluna = visual["coluna"]
    emoji = visual["emoji"]
    fill = visual["fill"]

    cel_manha = _localizar_celula_por_texto(ws, "MANHÃƒ")
    cel_tarde = _localizar_celula_por_texto(ws, "TARDE")
    cel_noite = _localizar_celula_por_texto(ws, "NOITE")
    cel_sol = _localizar_celula_por_texto(ws, "SOL")
    cel_nublado = _localizar_celula_por_texto(ws, "NUBLADO")
    cel_chuva = _localizar_celula_por_texto(ws, "CHUVA")

    mapa_col = {
        "SOL": cel_sol.column if cel_sol else None,
        "NUBLADO": cel_nublado.column if cel_nublado else None,
        "CHUVA": cel_chuva.column if cel_chuva else None,
    }
    col_destino = mapa_col.get(coluna)

    for cel_periodo in [cel_manha, cel_tarde]:
        if cel_periodo and col_destino:
            alvo = ws.cell(row=cel_periodo.row, column=col_destino)
            _marcar_celula_climatica(alvo, emoji, fill)

    if cel_noite and col_destino:
        alvo = ws.cell(row=cel_noite.row, column=col_destino)
        alvo.value = "€¢"
        _aplicar_estilo_celula(alvo, fill=fill, bold=True, size=12, align_center=True)


def preencher_bloco_horario(ws):
    cab = _localizar_celula_por_texto(ws, "HORÃRIO DE TRABALHO")
    if not cab:
        return

    linhas = [
        f"Início: {HORARIO_INICIO_PADRAO}",
        f"Intervalo: {HORARIO_INTERVALO_INICIO_PADRAO} Ã s {HORARIO_INTERVALO_FIM_PADRAO}",
        f"Término: {HORARIO_TERMINO_PADRAO}",
    ]
    for idx, texto in enumerate(linhas, start=1):
        destino = ws.cell(row=cab.row + idx, column=cab.column)
        destino.value = texto
        _aplicar_estilo_celula(destino, fill=FILL_INFO, size=9, wrap=True)


def preencher_header_limpo(ws, rdc, clima):
    supervisor_nome = rdc.supervisor.nome if getattr(rdc, "supervisor_id", None) else SUPERVISOR_PADRAO_NOME
    equipe_nome = _equipe_principal_nome(rdc)

    dias_semana_pt = {
        "Monday": "SEGUNDA-FEIRA",
        "Tuesday": "TERÃ‡A-FEIRA",
        "Wednesday": "QUARTA-FEIRA",
        "Thursday": "QUINTA-FEIRA",
        "Friday": "SEXTA-FEIRA",
        "Saturday": "SÃBADO",
        "Sunday": "DOMINGO",
    }
    dia_semana = dias_semana_pt.get(rdc.data.strftime("%A"), rdc.data.strftime("%A"))

    for cell_ref in ["G1", "G4", "H7", "M1", "M3", "M8", "M9"]:
        try:
            _resolver_celula_editavel(ws, cell_ref).value = None
        except Exception:
            pass

    titulo = _resolver_celula_editavel(ws, "G1")
    titulo.value = "RELATÃ“RIO DIÃRIO DE CAMPO - RDC"
    _aplicar_estilo_celula(
        titulo,
        fill=FILL_HEADER_BOX,
        bold=True,
        size=14,
        align_center=True,
        wrap=True,
    )

    bloco = _resolver_celula_editavel(ws, "G4")
    bloco.value = (
        f"Projeto: {rdc.projeto.codigo}\n"
        f"Disciplina: {rdc.disciplina.nome}\n"
        f"Área/local: {rdc.area_local.descricao}\n"
        f"Equipe: {equipe_nome}\n"
        f"Supervisor: {supervisor_nome}"
    )
    _aplicar_estilo_celula(
        bloco,
        fill=FILL_HEADER_BOX_2,
        bold=True,
        size=10,
        wrap=True,
    )

    cel_dia = _resolver_celula_editavel(ws, "M1")
    cel_dia.value = dia_semana
    _aplicar_estilo_celula(
        cel_dia,
        fill=FILL_INFO,
        bold=True,
        size=11,
        align_center=True,
        wrap=True,
    )

    cel_data = _resolver_celula_editavel(ws, "M3")
    cel_data.value = rdc.data.strftime("%d/%m/%Y")
    _aplicar_estilo_celula(
        cel_data,
        fill=FILL_INFO,
        bold=True,
        size=12,
        align_center=True,
        wrap=True,
    )

    cond_area = _resolver_celula_editavel(ws, "H7")
    cond_area.value = rdc.condicao_area or "LIBERADA"
    _aplicar_estilo_celula(
        cond_area,
        fill=FILL_INFO,
        bold=True,
        size=10,
        align_center=True,
        wrap=True,
    )

    status_area = _resolver_celula_editavel(ws, "M8")
    status_area.value = f"STATUS DA Área: {rdc.condicao_area or 'LIBERADA'}"
    _aplicar_estilo_celula(
        status_area,
        fill=FILL_INFO,
        bold=True,
        size=10,
        align_center=True,
        wrap=True,
    )

    resumo = _resolver_celula_editavel(ws, "M9")
    if clima.get("temperatura_min") is not None and clima.get("temperatura_max") is not None:
        resumo.value = (
            f"CLIMA: {clima.get('descricao')} | "
            f"TEMP: {clima.get('temperatura_min')}°C / {clima.get('temperatura_max')}°C | "
            f"HORÃRIO: {HORARIO_INICIO_PADRAO}-{HORARIO_TERMINO_PADRAO}"
        )
    else:
        resumo.value = (
            f"CLIMA: {clima.get('descricao')} | "
            f"HORÃRIO: {HORARIO_INICIO_PADRAO}-{HORARIO_TERMINO_PADRAO}"
        )
    _aplicar_estilo_celula(
        resumo,
        fill=FILL_INFO,
        bold=True,
        size=9,
        wrap=True,
    )


def destacar_supervisor_no_quadro(ws):
    for linha in range(13, 21):
        cargo = _normalizar_texto(ws[f"I{linha}"].value)
        if "SUPERVISOR" in cargo:
            for col in ["A", "B", "I", "K", "L", "M", "N", "O", "P"]:
                _aplicar_estilo_celula(ws[f"{col}{linha}"], fill=FILL_SUPERVISOR, bold=True, wrap=True)
            break


def aplicar_estilo_atividades(ws, atividade_linhas):
    for linha in atividade_linhas:
        _aplicar_estilo_celula(ws[f"A{linha}"], fill=FILL_TITULO_FORTE, bold=True, align_center=True)
        _aplicar_estilo_celula(ws[f"B{linha}"], fill=FILL_DESTAQUE, wrap=True)
        _aplicar_estilo_celula(ws[f"N{linha}"], fill=FILL_INFO, wrap=True)


def aplicar_estilo_hh(ws, atividade_colunas):
    for coluna in atividade_colunas:
        _aplicar_estilo_celula(ws[f"{coluna}21"], fill=FILL_CABECALHO, bold=True, align_center=True)


def aplicar_estilo_quadro_equipe_elite(ws):
    for linha in range(13, 21):
        cargo = _normalizar_texto(ws[f"I{linha}"].value)
        fill_linha = FILL_LIDERANCA if _funcao_eh_lideranca(cargo) else FILL_EQUIPE
        bold = _funcao_eh_lideranca(cargo)

        for col in ["A", "B", "I", "K", "L", "M", "N", "O", "P"]:
            _aplicar_estilo_celula(ws[f"{col}{linha}"], fill=fill_linha, bold=bold, wrap=True)

        ws[f"A{linha}"].alignment = Alignment(horizontal="center", vertical="center")
        for col in ["K", "L", "M", "N", "O", "P"]:
            ws[f"{col}{linha}"].alignment = Alignment(horizontal="center", vertical="center")


def preencher_bloco_observacoes_elite(ws, rdc, clima):
    observacoes = [
        f"Obra: {CIDADE_OBRA}",
        f"Projeto: {rdc.projeto.codigo}",
        f"Disciplina: {rdc.disciplina.nome}",
        f"Área/local: {rdc.area_local.descricao}",
        f"Supervisor: {rdc.supervisor.nome if getattr(rdc, 'supervisor_id', None) else SUPERVISOR_PADRAO_NOME}",
        f"Clima: {clima.get('descricao')}",
    ]
    ws["A35"] = f"COMENTÃRIOS EXECUTIVOS: {' | '.join(observacoes)}"
    _aplicar_estilo_celula(ws["A35"], fill=FILL_RODAPE, bold=True, size=9, wrap=True)


def preencher_bloco_restricoes_elite(ws, rdc):
    restricoes = []
    if rdc.validacoes.exists():
        total = rdc.validacoes.count()
        alertas = rdc.validacoes.filter(status=StatusValidacaoChoices.ALERTA).count()
        bloqueios = rdc.validacoes.filter(status=StatusValidacaoChoices.BLOQUEIO).count()
        infos = rdc.validacoes.filter(status=StatusValidacaoChoices.INFO).count()
        restricoes.extend([
            f"ValidaçÃµes registradas: {total}",
            f"Alertas: {alertas}",
            f"Bloqueios: {bloqueios}",
            f"InformaçÃµes: {infos}",
        ])
    else:
        restricoes.append("Sem validaçÃµes registradas")

    ws["A36"] = f"RESTRIÃ‡Ã•ES / QUALIDADE DE DADOS: {' | '.join(restricoes)}"
    _aplicar_estilo_celula(ws["A36"], fill=FILL_ALERTA, bold=True, size=9, wrap=True)


def preencher_bloco_assinatura_elite(ws, rdc):
    assinatura = [
        f"Emitido automaticamente em {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
        f"Responsável pela gerAção: {getattr(getattr(rdc, 'criado_por', None), 'username', 'sistema')}",
        f"Supervisor de referência: {rdc.supervisor.nome if getattr(rdc, 'supervisor_id', None) else SUPERVISOR_PADRAO_NOME}",
    ]
    ws["A37"] = "ASSINATURA / EMISSÃƒO: " + " | ".join(assinatura)
    _aplicar_estilo_celula(ws["A37"], fill=FILL_INFO, bold=True, size=9, wrap=True)


def exportar_rdc_para_modelo_excel(rdc):
    template_path = Path(
        getattr(settings, "RDC_TEMPLATE_PATH", settings.BASE_DIR.parent / "RDC - MODELO.xlsx")
    )
    if not template_path.exists():
        raise FileNotFoundError(f"Modelo RDC não encontrado em: {template_path}")

    wb = load_workbook(template_path)
    ws = wb["Planilha1"] if "Planilha1" in wb.sheetnames else wb.active

    if "Plan1" in wb.sheetnames and "Planilha1" in wb.sheetnames:
        wb["Plan1"].sheet_state = "hidden"

    clima = buscar_clima_online_rio_grande(rdc.data)

    ajustar_layout_topo(ws)

    logo_ok = inserir_logo(ws)
    if not logo_ok:
        print("[RDC] ExportAção seguirá sem logo.")

    preencher_quadro_climatico(ws, clima)
    preencher_bloco_horario(ws)
    preencher_header_limpo(ws, rdc, clima)

    atividade_colunas = ["K", "L", "M", "N", "O", "P"]
    atividade_linhas = [23, 25, 27, 29, 31, 33]
    letras_atividades = ["A", "B", "C", "D", "E", "F"]
    atividades = selecionar_atividades_unicas_para_exportacao(rdc, limite=6)

    for idx, linha in enumerate(atividade_linhas):
        atividade = atividades[idx] if idx < len(atividades) else None
        ws[f"A{linha}"] = letras_atividades[idx]

        if atividade:
            partes = []
            if atividade.codigo_atividade:
                partes.append(str(atividade.codigo_atividade).strip())
            if atividade.descr_atividade:
                partes.append(str(atividade.descr_atividade).strip())
            if atividade.codigo_subatividade:
                partes.append(f"/ {str(atividade.codigo_subatividade).strip()}")
            if atividade.descr_subatividade:
                partes.append(f"- {str(atividade.descr_subatividade).strip()}")

            ws[f"B{linha}"] = " ".join(partes).strip()
            ws[f"N{linha}"] = rdc.area_local.descricao
        else:
            ws[f"B{linha}"] = None
            ws[f"N{linha}"] = None

    aplicar_estilo_atividades(ws, atividade_linhas)

    funcionarios_rdc = list(rdc.funcionarios.select_related("funcao", "equipe", "funcionario").all())
    funcionarios_rdc.sort(
        key=lambda x: (
            0 if getattr(getattr(x, "equipe", None), "codigo", "") == EQUIPE_TESTE_PRIORITARIA else 1,
            _peso_funcao(getattr(getattr(x, "funcao", None), "nome", "")),
            (x.nome or "").upper(),
        )
    )
    funcionarios_rdc = funcionarios_rdc[:8]
    funcionarios = funcionarios_rdc or buscar_funcionarios_reais_por_catraca_para_rdc(rdc.data, limite=8)

    apontamentos = list(rdc.apontamentos.select_related("rdc_funcionario", "rdc_atividade").all())
    mapa_atividade = {atividade.id: letras_atividades[idx] for idx, atividade in enumerate(atividades)}
    horas_por_letra = {letra: Decimal("0.00") for letra in letras_atividades}
    letras_por_funcionario = defaultdict(list)

    for apontamento in apontamentos:
        letra = mapa_atividade.get(apontamento.rdc_atividade_id)
        if not letra:
            continue
        horas_por_letra[letra] += apontamento.horas or Decimal("0.00")
        letras_por_funcionario[apontamento.rdc_funcionario_id].append(letra)

    for indice in range(8):
        linha = 13 + indice
        funcionario = funcionarios[indice] if indice < len(funcionarios) else None

        ws[f"A{linha}"] = indice + 1
        _aplicar_estilo_celula(ws[f"A{linha}"], fill=FILL_CABECALHO, bold=True, align_center=True)

        if funcionario:
            ws[f"B{linha}"] = getattr(funcionario, "nome", None)
            ws[f"I{linha}"] = getattr(getattr(funcionario, "funcao", None), "nome", None)
        else:
            ws[f"B{linha}"] = None
            ws[f"I{linha}"] = None

        _aplicar_estilo_celula(ws[f"B{linha}"], fill=FILL_DESTAQUE, wrap=True)
        _aplicar_estilo_celula(ws[f"I{linha}"], fill=FILL_DESTAQUE, wrap=True)

        for coluna in atividade_colunas:
            ws[f"{coluna}{linha}"] = None
            _aplicar_estilo_celula(ws[f"{coluna}{linha}"], align_center=True)

        if not funcionario:
            continue

        funcionario_id = getattr(funcionario, "id", None)
        letras = letras_por_funcionario.get(funcionario_id, []) if funcionario_id else []

        if letras:
            for pos, letra in enumerate(letras[:6]):
                ws[f"{atividade_colunas[pos]}{linha}"] = letra
        else:
            distribuicao = distribuir_horas_por_atividades(funcionario, atividades)
            for pos, (letra, horas) in enumerate(distribuicao[:6]):
                ws[f"{atividade_colunas[pos]}{linha}"] = letra
                horas_por_letra[letra] += horas

    destacar_supervisor_no_quadro(ws)
    aplicar_estilo_quadro_equipe_elite(ws)

    for coluna, letra in zip(atividade_colunas, letras_atividades):
        valor = float(horas_por_letra.get(letra, Decimal("0.00")))
        ws[f"{coluna}21"] = valor if valor else None

    aplicar_estilo_hh(ws, atividade_colunas)
    preencher_bloco_observacoes_elite(ws, rdc, clima)
    preencher_bloco_restricoes_elite(ws, rdc)
    preencher_bloco_assinatura_elite(ws, rdc)

    export_dir = Path(settings.MEDIA_ROOT) / "rdc_exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%H%M%S")
    out_path = export_dir / f"rdc_{rdc.id}_{rdc.data.strftime('%Y%m%d')}_{timestamp}.xlsx"
    wb.save(out_path)

    print(f"[RDC] Arquivo exportado: {out_path}")
    return out_path

















from .rdc_contexto_service import (
    buscar_programacao_semana,
)









from .rdc_montagem_service import (
    sugerir_contexto_rdc_por_cronograma,
    gerar_apontamentos_base_automaticos,
    montar_rdc_pre_preenchido,
    montar_rdc_simulado_por_cronograma,
)
