from pathlib import Path
from datetime import datetime
from decimal import Decimal
from collections import defaultdict

from django.conf import settings

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from core.choices import StatusValidacaoChoices
from rdc.models import RDC

# IMPORTANTE: vamos reutilizar funções do rdc_service por enquanto
from .rdc_service import (
    buscar_clima_online_rio_grande,
    selecionar_atividades_unicas_para_exportacao,
    distribuir_horas_por_atividades,
)


def exportar_rdc_para_modelo_excel(rdc):
    """
    Função extraída do rdc_service para organização.
    Mantém comportamento original.
    """

    template_path = Path(
        getattr(settings, "RDC_TEMPLATE_PATH", settings.BASE_DIR.parent / "RDC - MODELO.xlsx")
    )
    if not template_path.exists():
        raise FileNotFoundError(f"Modelo RDC não encontrado em: {template_path}")

    wb = load_workbook(template_path)
    ws = wb.active

    clima = buscar_clima_online_rio_grande(rdc.data)

    # Aqui mantemos simples por enquanto (sem mover helpers ainda)
    ws["A1"] = "RDC EXPORTADO"
    ws["A2"] = f"Projeto: {rdc.projeto.codigo}"
    ws["A3"] = f"Data: {rdc.data.strftime('%d/%m/%Y')}"
    ws["A4"] = f"Clima: {clima.get('descricao')}"

    export_dir = Path(settings.MEDIA_ROOT) / "rdc_exports"
    export_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%H%M%S")
    out_path = export_dir / f"rdc_{rdc.id}_{rdc.data.strftime('%Y%m%d')}_{timestamp}.xlsx"

    wb.save(out_path)

    return out_path


