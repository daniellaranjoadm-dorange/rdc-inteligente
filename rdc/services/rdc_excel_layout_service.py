from decimal import Decimal
from pathlib import Path

from django.conf import settings

from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


CIDADE_OBRA = "Rio Grande/RS - Estaleiro Rio Grande"
LATITUDE_OBRA = -32.035
LONGITUDE_OBRA = -52.098
TIMEZONE_OBRA = "America/Sao_Paulo"

HORARIO_INICIO_PADRAO = "07:30"
HORARIO_INTERVALO_INICIO_PADRAO = "12:00"
HORARIO_INTERVALO_FIM_PADRAO = "13:00"
HORARIO_TERMINO_PADRAO = "17:30"

SUPERVISOR_PADRAO_NOME = "Daniel Laranjo"
EQUIPE_TESTE_PRIORITARIA = "RDC-TST"

RDC_LOGO_FILENAMES = [
    "Logo MARENOVA_novo.PNG",
    "Logo MARENOVA_novo.png",
    "logo_marenova.png",
    "logo_marenova.PNG",
    "logo.png",
    "Logo.png",
]

FILL_SOL = PatternFill(fill_type="solid", fgColor="FFF2CC")
FILL_NUBLADO = PatternFill(fill_type="solid", fgColor="D9E2F3")
FILL_CHUVA = PatternFill(fill_type="solid", fgColor="BDD7EE")
FILL_INFO = PatternFill(fill_type="solid", fgColor="E2F0D9")
FILL_ALERTA = PatternFill(fill_type="solid", fgColor="FCE4D6")
FILL_DESTAQUE = PatternFill(fill_type="solid", fgColor="DDEBF7")
FILL_CABECALHO = PatternFill(fill_type="solid", fgColor="D9EAD3")
FILL_SUPERVISOR = PatternFill(fill_type="solid", fgColor="FFF2CC")
FILL_EQUIPE = PatternFill(fill_type="solid", fgColor="F3F6FA")
FILL_LIDERANCA = PatternFill(fill_type="solid", fgColor="FFF2CC")
FILL_RODAPE = PatternFill(fill_type="solid", fgColor="EAF2F8")
FILL_TITULO_FORTE = PatternFill(fill_type="solid", fgColor="B4C6E7")
FILL_HEADER_BOX = PatternFill(fill_type="solid", fgColor="DDEBF7")
FILL_HEADER_BOX_2 = PatternFill(fill_type="solid", fgColor="E2F0D9")

BORDER_THIN = Border(
    left=Side(style="thin", color="999999"),
    right=Side(style="thin", color="999999"),
    top=Side(style="thin", color="999999"),
    bottom=Side(style="thin", color="999999"),
)


def _normalizar_texto(valor):
    return str(valor or "").strip().upper()


def _resolver_celula_editavel(ws, cell_ref):
    cell = ws[cell_ref]
    if not isinstance(cell, MergedCell):
        return cell

    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col)

    return cell


def _aplicar_estilo_celula(cell, fill=None, bold=False, size=None, align_center=False, wrap=False):
    if fill:
        cell.fill = fill
    cell.border = BORDER_THIN
    cell.alignment = Alignment(
        horizontal="center" if align_center else "left",
        vertical="center",
        wrap_text=wrap,
    )
    cell.font = Font(bold=bold, size=size if size else 10)


def _marcar_celula_climatica(cell, emoji, fill):
    cell.value = emoji
    _aplicar_estilo_celula(cell, fill=fill, bold=True, size=13, align_center=True)


def _resolver_logo_path():
    caminhos = []

    logo_setting = getattr(settings, "RDC_LOGO_PATH", None)
    if logo_setting:
        caminhos.append(Path(logo_setting))

    base_dir = Path(settings.BASE_DIR)
    base_parent = base_dir.parent
    arquivo_atual_dir = Path(__file__).resolve().parent
    projeto_raiz = arquivo_atual_dir.parent

    bases = [
        base_dir,
        base_parent,
        projeto_raiz,
        arquivo_atual_dir,
        base_dir / "static",
        base_dir / "static" / "img",
        base_parent / "static",
        base_parent / "static" / "img",
        projeto_raiz / "static",
        projeto_raiz / "static" / "img",
    ]

    for base in bases:
        for nome in RDC_LOGO_FILENAMES:
            caminhos.append(base / nome)

    for caminho in caminhos:
        try:
            if caminho and Path(caminho).exists() and Path(caminho).is_file():
                return Path(caminho)
        except Exception:
            continue

    return None


def inserir_logo(ws):
    logo_path = _resolver_logo_path()
    if not logo_path:
        print("[RDC] Logo não encontrada. Verifique RDC_LOGO_PATH ou coloque a imagem em static/img.")
        return False

    try:
        img = XLImage(str(logo_path))
        img.width = 185
        img.height = 78
        ws.add_image(img, "A1")
        print(f"[RDC] Logo inserida com sucesso: {logo_path}")
        return True
    except Exception as exc:
        print(f"[RDC] Falha ao inserir logo '{logo_path}': {exc}")
        return False


def ajustar_layout_topo(ws):
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 24
    ws.row_dimensions[4].height = 24
    ws.row_dimensions[5].height = 28
    ws.row_dimensions[7].height = 22
    ws.row_dimensions[8].height = 22
    ws.row_dimensions[9].height = 24
    ws.row_dimensions[10].height = 24

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 18
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 8
    ws.column_dimensions["L"].width = 8
    ws.column_dimensions["M"].width = 16
    ws.column_dimensions["N"].width = 16
    ws.column_dimensions["O"].width = 16
    ws.column_dimensions["P"].width = 16


