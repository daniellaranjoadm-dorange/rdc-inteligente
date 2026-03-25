
from decimal import Decimal
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from cadastros.models import AreaLocal, Disciplina, Equipe, Funcionario, Projeto
from planejamento.models import AtividadeCronograma
from rdc.models import CalendarioPlanejamento, ProgramacaoSemanal


class Command(BaseCommand):
    help = 'Importa programAção semanal a partir de planilha.'

    def add_arguments(self, parser):
        parser.add_argument('--arquivo', required=True)
        parser.add_argument('--projeto-codigo', required=True)
        parser.add_argument('--aba', default=None)
        parser.add_argument('--limpar', action='store_true')

    def handle(self, *args, **options):
        arquivo = Path(options['arquivo'])
        if not arquivo.exists():
            raise CommandError(f'Arquivo não encontrado: {arquivo}')
        projeto = Projeto.objects.filter(codigo=options['projeto_codigo']).first()
        if not projeto:
            raise CommandError('Projeto não encontrado pelo código informado.')

        wb = load_workbook(arquivo, data_only=True)
        ws = wb[options['aba']] if options.get('aba') else wb[wb.sheetnames[0]]

        header_row = None
        header_map = {}
        expected = {
            'semana': ['semana', 'semana_codigo'],
            'data': ['data', 'data_programada'],
            'disciplina': ['disciplina'],
            'area': ['area', 'área', 'area_local'],
            'equipe': ['equipe'],
            'encarregado': ['encarregado'],
            'codigo_atividade': ['codigo_atividade', 'código', 'codigo'],
            'descr_atividade': ['descr_atividade', 'descrição', 'descricao'],
            'codigo_subatividade': ['codigo_subatividade', 'subcodigo', 'subcódigo'],
            'descr_subatividade': ['descr_subatividade', 'subdescrição', 'subdescricao'],
            'qtd_prevista': ['qtd_prevista', 'quantidade', 'qtd'],
            'hh_previsto': ['hh_previsto', 'hh'],
            'turno': ['turno'],
            'observacao': ['observacao', 'observAção'],
        }

        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
            normalized = [str(c).strip().lower() if c is not None else '' for c in row]
            found = {}
            for idx, value in enumerate(normalized):
                for key, aliases in expected.items():
                    if value in aliases:
                        found[key] = idx
            if 'semana' in found and 'codigo_atividade' in found and 'descr_atividade' in found:
                header_row = i
                header_map = found
                break
        if not header_row:
            raise CommandError('Não foi possível localizar o cabeçalho da programAção semanal na planilha.')

        if options['limpar']:
            ProgramacaoSemanal.objects.filter(projeto=projeto).delete()

        created = 0
        with transaction.atomic():
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                semana_codigo = row[header_map['semana']] if header_map.get('semana') is not None else None
                codigo_atividade = row[header_map['codigo_atividade']] if header_map.get('codigo_atividade') is not None else None
                descr_atividade = row[header_map['descr_atividade']] if header_map.get('descr_atividade') is not None else None
                if not semana_codigo or not codigo_atividade or not descr_atividade:
                    continue
                semana_codigo = str(semana_codigo).strip()
                data_programada = row[header_map['data']] if header_map.get('data') is not None else None
                disciplina = None
                area_local = None
                equipe = None
                encarregado = None
                atividade_cronograma = None
                if header_map.get('disciplina') is not None and row[header_map['disciplina']]:
                    disciplina = Disciplina.objects.filter(nome__iexact=str(row[header_map['disciplina']]).strip()).first()
                if header_map.get('area') is not None and row[header_map['area']]:
                    area_local = AreaLocal.objects.filter(projeto=projeto, descricao__iexact=str(row[header_map['area']]).strip()).first()
                if header_map.get('equipe') is not None and row[header_map['equipe']]:
                    equipe = Equipe.objects.filter(nome__iexact=str(row[header_map['equipe']]).strip()).first()
                if header_map.get('encarregado') is not None and row[header_map['encarregado']]:
                    encarregado = Funcionario.objects.filter(nome__iexact=str(row[header_map['encarregado']]).strip()).first()
                atividade_cronograma = AtividadeCronograma.objects.filter(projeto=projeto, codigo_atividade=str(codigo_atividade).strip()).first()
                calendario = CalendarioPlanejamento.objects.filter(projeto=projeto, semana_codigo=semana_codigo).order_by('data').first()
                ProgramacaoSemanal.objects.create(
                    projeto=projeto,
                    semana_codigo=semana_codigo,
                    semana_label=f'SEM {semana_codigo}',
                    data_inicio_semana=getattr(calendario, 'data_inicio_semana', None),
                    data_fim_semana=getattr(calendario, 'data_fim_semana', None),
                    data_programada=data_programada if hasattr(data_programada, 'year') else None,
                    disciplina=disciplina,
                    area_local=area_local,
                    equipe=equipe,
                    encarregado=encarregado,
                    atividade_cronograma=atividade_cronograma,
                    codigo_atividade=str(codigo_atividade).strip(),
                    descr_atividade=str(descr_atividade).strip(),
                    codigo_subatividade=str(row[header_map['codigo_subatividade']]).strip() if header_map.get('codigo_subatividade') is not None and row[header_map['codigo_subatividade']] else '',
                    descr_subatividade=str(row[header_map['descr_subatividade']]).strip() if header_map.get('descr_subatividade') is not None and row[header_map['descr_subatividade']] else '',
                    qtd_prevista=Decimal(str(row[header_map['qtd_prevista']]).replace(',', '.')) if header_map.get('qtd_prevista') is not None and row[header_map['qtd_prevista']] not in (None, '') else Decimal('0.00'),
                    hh_previsto=Decimal(str(row[header_map['hh_previsto']]).replace(',', '.')) if header_map.get('hh_previsto') is not None and row[header_map['hh_previsto']] not in (None, '') else Decimal('0.00'),
                    turno=str(row[header_map['turno']]).strip() if header_map.get('turno') is not None and row[header_map['turno']] else 'integral',
                    observacao=str(row[header_map['observacao']]).strip() if header_map.get('observacao') is not None and row[header_map['observacao']] else '',
                    origem_arquivo=arquivo.name,
                )
                created += 1
        self.stdout.write(self.style.SUCCESS(f'ProgramAção semanal importada com sucesso. Registros: {created}'))


