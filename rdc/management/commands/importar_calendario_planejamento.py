from datetime import timedelta
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from cadastros.models import Projeto
from rdc.models import CalendarioPlanejamento


class Command(BaseCommand):
    help = "Importa o calendário de planejamento (semanas + eventos) para uso no RDC guiado."

    def add_arguments(self, parser):
        parser.add_argument("--arquivo", required=True, help="Caminho do arquivo XLSX do calendário.")
        parser.add_argument("--projeto-id", type=int, help="ID do projeto.")
        parser.add_argument("--projeto-codigo", help="Código do projeto.")
        parser.add_argument("--limpar", action="store_true", help="Apaga o calendário existente do projeto antes de importar.")

    def handle(self, *args, **options):
        arquivo = Path(options["arquivo"])
        if not arquivo.exists():
            raise CommandError(f"Arquivo não encontrado: {arquivo}")

        projeto = None
        if options.get("projeto_id"):
            projeto = Projeto.objects.filter(pk=options["projeto_id"]).first()
        elif options.get("projeto_codigo"):
            projeto = Projeto.objects.filter(codigo__iexact=options["projeto_codigo"]).first()

        if projeto is None:
            raise CommandError("Informe --projeto-id ou --projeto-codigo com um projeto válido.")

        wb = load_workbook(arquivo, data_only=True)
        ws = wb.active

        codigos = set()
        for row in ws.iter_rows():
            for cell in row:
                valor = cell.value
                if isinstance(valor, (int, float)):
                    codigo = str(int(valor))
                    if len(codigo) == 4 and codigo.isdigit():
                        ano = int(codigo[:2])
                        semana = int(codigo[2:])
                        if 20 <= ano <= 40 and 1 <= semana <= 53:
                            codigos.add(codigo)

        eventos = {}
        for row in range(34, ws.max_row + 1):
            for col in range(1, ws.max_column - 2):
                data_evento = ws.cell(row, col).value
                descricao = ws.cell(row, col + 2).value
                if hasattr(data_evento, "date") and descricao:
                    eventos[data_evento.date()] = str(descricao).strip()

        if options["limpar"]:
            CalendarioPlanejamento.objects.filter(projeto=projeto).delete()

        total_registros = 0
        for codigo in sorted(codigos):
            ano = 2000 + int(codigo[:2])
            semana_numero = int(codigo[2:])
            jan1 = __import__("datetime").date(ano, 1, 1)
            primeiro_sabado = jan1 + timedelta(days=(5 - jan1.weekday()) % 7)
            inicio_semana = primeiro_sabado + timedelta(weeks=semana_numero - 1)
            fim_semana = inicio_semana + timedelta(days=6)

            for offset, (weekday_num, dia_nome) in enumerate([
                (5, "Sábado"),
                (6, "Domingo"),
                (0, "Segunda-feira"),
                (1, "Terça-feira"),
                (2, "Quarta-feira"),
                (3, "Quinta-feira"),
                (4, "Sexta-feira"),
            ]):
                data_ref = inicio_semana + timedelta(days=offset)
                descricao_evento = eventos.get(data_ref, "")
                eh_feriado = "FERIADO" in descricao_evento.upper() if descricao_evento else False
                eh_dia_util = weekday_num not in (5, 6) and not eh_feriado

                CalendarioPlanejamento.objects.update_or_create(
                    projeto=projeto,
                    data=data_ref,
                    defaults={
                        "ano": data_ref.year,
                        "mes": data_ref.month,
                        "semana_codigo": codigo,
                        "semana_numero": semana_numero,
                        "semana_label": f"SEM {codigo}",
                        "data_inicio_semana": inicio_semana,
                        "data_fim_semana": fim_semana,
                        "dia_semana": weekday_num,
                        "dia_semana_nome": dia_nome,
                        "eh_dia_util": eh_dia_util,
                        "eh_feriado": eh_feriado,
                        "descricao_evento": descricao_evento,
                    },
                )
                total_registros += 1

        self.stdout.write(self.style.SUCCESS(
            f"Calendário importado com sucesso para {projeto.codigo}. Semanas: {len(codigos)} | registros/dias: {total_registros}"
        ))

